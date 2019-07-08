"""
Created on MON 0204 22:04:46 2019
@author: ChihChi
"""
import pandas as pd
import numpy as np
import random
import math
import time
from openpyxl import Workbook
from openpyxl import load_workbook


# 參數設定
NumOfPerson = 20
Hours = 24
E_s_min = 0
E_ev_min = 12.5
E_s_max = 13.5
E_ev_max = 50
Rate_s = 5
Rate_ev = 10 
Beta_s = 0.1
Beta_ev = 0.1

# 目前週幾
# DAY = 0 -> Monday, DAY = 1 -> Tuesday, DAY = 2 -> Wed
DAY = 5
DAY_List = [ 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']


time_start = time.time()

Hist_BestFit = list()
Hist_AvgFit = list()

# 演算法參數
MaxOfIteration = 25000

Par_a = 0
Par_A = 0
Par_C = 0
Par_l = 0
Par_p = 0
Par_b = 1       # 對數螺旋

a = random.randint( 1, 150000)
a = 106076


# 設定亂數種子
random.seed( a)
np.random.seed( a)


# 建立物件
class person:
    def __init__(self):
        self.E_uug = list()
        self.E_sfg = list()
        self.E_uus = list()
        self.E_sso = list()
        self.E_sfo = list()
        self.E_ubo = list()
        self.E_vfg = list()
        self.E_uuv = list()
        self.E_vfs = list()
        self.E_vso = list()
        self.E_vfo = list()
        self.theta_s = list()
        self.theta_ev = list()
        self.check_s_on_t = list()
        self.check_ev_on_t = list()
        self.fitness = 0
        
# =============================================================================        
PathOfFile = 'C:/Users/juice/Desktop/MyProposal/IoE_Program/'

# == 可再生能源 ==
target = pd.read_excel( PathOfFile + 'RenewableEnergy.xlsx', dtype = 'float64')
ParaOfRenergy = target.values

# == 實時電價 ==
target = pd.read_excel( PathOfFile + 'RTP.xlsx', dtype = 'float64')
ParaOfRTP = target.values

# == 市場價格 ==
target = pd.read_excel( PathOfFile + 'MarketingPrice.xlsx', dtype = 'float64')
ParaOfMarketP = target.values

# == D_user ==
target = pd.read_excel( PathOfFile + 'Duser.xlsx', dtype = 'float64')
D_user = target.values

# == Theta_Work ==
target = pd.read_excel( PathOfFile + 'Theta_work.xlsx', dtype = 'float64')
Theta_Work = target.values

# =============================================================================

# function
def calTheFit( person):
    
    C_grid, C_user, R_profit, R_capital = 0, 0, 0, 0
    for hour_i in range( Hours):
        C_grid = C_grid + ParaOfRTP[DAY][hour_i] * ( person.E_uug[hour_i] + ( person.E_sfg[hour_i] / ( 1 - Beta_s)) + ( person.E_vfg[hour_i] / (1 - Beta_ev)))
        C_user = C_user + ParaOfMarketP[DAY][hour_i] * ( person.E_ubo[hour_i] + ( person.E_sfo[hour_i] / ( 1- Beta_s)) + ( person.E_vfo[hour_i] / (1 - Beta_ev)))
        R_profit = R_profit + ParaOfRTP[DAY][hour_i] * ( person.E_uus[hour_i] + person.E_uuv[hour_i]) + ParaOfMarketP[DAY][hour_i] * (person.E_sso[hour_i] + person.E_vso[hour_i])
    R_capital = (( ParaOfMarketP.sum(axis = 1)[DAY]) + ( ParaOfRTP.sum(axis = 1)[DAY])) * (person.check_s_on_t[-1] + person.check_ev_on_t[-1]) / 48
    Fitness = C_grid + C_user - R_profit - R_capital

    return Fitness

def ChechLegalOrNot( person):
    if min(person.check_s_on_t) < 0:
        print( "Storage illegal")
    if min(person.check_ev_on_t) < E_ev_min:
        print( "EV illegal")




def CycleMove( now_w, best_w):
    
    global Par_b
    global Par_l
    
    new_location = max( abs( now_w - best_w) * math.exp( Par_b * Par_l) * math.cos( 2 * math.pi * Par_l ) + now_w , 0)
    return (new_location)






# Step1:產生起始解    
population = list()
for i in range( NumOfPerson):
    population.append( person())
    
# 起始值
for i in range( NumOfPerson):
    
    # 產生人口前先初始化 
    tmp_E_uug, tmp_E_sfg, tmp_E_uus = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_E_sso, tmp_E_sfo, tmp_E_ubo = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_E_vfg, tmp_E_uuv, tmp_E_vfs = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_E_vso, tmp_E_vfo, tmp_theta_s, tmp_theta_ev = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_s_on_t, tmp_ev_on_t = np.zeros((1, Hours + 1)), np.zeros((1, Hours + 1))
    
    if DAY == 0:
        tmp_s_on_t[0][0] = E_s_min
        tmp_ev_on_t[0][0] = E_ev_min
    else:
        file = load_workbook( PathOfFile + 'Remain0416.xlsx')
        sheet_ranges = file['HowManyStillHave_WOA']
    
        for k in range( len( DAY_List)):
            if DAY == list(enumerate(DAY_List))[k][0]:
                tmp_s_on_t[0][0] = (sheet_ranges.cell( row = k + 1, column = 2).value)
                tmp_ev_on_t[0][0] = (sheet_ranges.cell( row = k + 1, column = 3).value)
                break
    
    ######
    excel_s_on_t = tmp_s_on_t[0][0].copy()
    excel_ev_on_t = tmp_ev_on_t[0][0].copy()
    ######
    
    population[i].check_s_on_t.append( tmp_s_on_t[0][0])
    population[i].check_ev_on_t.append( tmp_ev_on_t[0][0])
    
    for hour_i in range( Hours):
        
        # 確認儲電設備
        if tmp_s_on_t[0][hour_i] <= E_s_min:                  # 儲電設備沒存量不能放電
            tmp_theta_s[0][hour_i] = 1
        elif tmp_s_on_t[0][hour_i] + ParaOfRenergy[DAY][hour_i] >= E_s_max:   # 儲電設備沒容量不能儲電
            tmp_theta_s[0][hour_i] = 0
        else:
            tmp_theta_s[0][hour_i] = random.randint(0, 1)
        
        # 確認電動
        if tmp_ev_on_t[0][hour_i] <= E_ev_min:
            tmp_theta_ev[0][hour_i] = 1
        elif tmp_ev_on_t[0][hour_i] >= E_ev_max or Theta_Work[DAY][hour_i] == 1:
            tmp_theta_ev[0][hour_i] = 0
        else:
            tmp_theta_ev[0][hour_i] = random.randint(0, 1)
        
        # 產生儲放電值
        
        # 當儲電設備儲電，電動車儲電
        if tmp_theta_s[0][hour_i] == 1 and tmp_theta_ev[0][hour_i] == 1:            
            
            # 儲電設備
            tmp_E_sfg[0][hour_i] = random.uniform(0, min( (E_s_max - tmp_s_on_t[0][hour_i] - ParaOfRenergy[DAY][hour_i]), Rate_s))
            tmp_E_sfo[0][hour_i] = random.uniform(0, min( max(E_s_max - tmp_s_on_t[0][hour_i] - tmp_E_sfg[0][hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s))        
            
            #電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_vfg[0][hour_i] = random.uniform(0, min(E_ev_max - tmp_ev_on_t[0][hour_i], Rate_ev))
                tmp_E_vfo[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfg[0][hour_i], 0), Rate_ev))
        
        # 當儲電設備放電，電動車儲電
        elif tmp_theta_s[0][hour_i] == 0 and tmp_theta_ev[0][hour_i] == 1:
            
            # 儲電設備
            tmp_E_uus[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min( tmp_s_on_t[0][hour_i], Rate_s)))
            tmp_E_sso[0][hour_i] = random.uniform(0, min(max( tmp_s_on_t[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s))
            
            #電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_vfs[0][hour_i] = random.uniform(0, min(E_ev_max - tmp_ev_on_t[0][hour_i], min(max( tmp_s_on_t[0][hour_i] - tmp_E_sso[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s)))
                tmp_E_vfg[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfs[0][hour_i],0), Rate_ev))
                tmp_E_vfo[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfg[0][hour_i] - tmp_E_vfs[0][hour_i], 0), Rate_ev))
        
        # 當儲電設備儲電，電動車放電
        elif tmp_theta_s[0][hour_i] == 1 and tmp_theta_ev[0][hour_i] == 0:
        
            # 儲電設備
            tmp_E_sfg[0][hour_i] = random.uniform(0, min( (E_s_max - tmp_s_on_t[0][hour_i] - ParaOfRenergy[DAY][hour_i]), Rate_s))
            tmp_E_sfo[0][hour_i] = random.uniform(0, min( max(E_s_max - tmp_s_on_t[0][hour_i] - tmp_E_sfg[0][hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s))
            
            # 電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_uuv[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min(tmp_ev_on_t[0][hour_i] - E_ev_min, Rate_ev)))
                tmp_E_vso[0][hour_i] = random.uniform(0, min(max(tmp_ev_on_t[0][hour_i] - tmp_E_uuv[0][hour_i] - E_ev_min, 0), Rate_ev))
        
        # 當儲電設備放電，電動車放電
        elif tmp_theta_s[0][hour_i] == 0 and tmp_theta_ev[0][hour_i] == 0:
        
            # 儲電設備
            tmp_E_uus[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min( tmp_s_on_t[0][hour_i], Rate_s)))
            tmp_E_sso[0][hour_i] = random.uniform(0, min(max( tmp_s_on_t[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s))
            # 電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_uuv[0][hour_i] = random.uniform(0, min(max(0, D_user[0][hour_i] - tmp_E_uus[0][hour_i]), min(tmp_ev_on_t[0][hour_i] - E_ev_min, Rate_ev)))
                tmp_E_vso[0][hour_i] = random.uniform(0, min(max(tmp_ev_on_t[0][hour_i] - tmp_E_uuv[0][hour_i] - E_ev_min, 0), Rate_ev))
        else:
            print('Error in initial!!')
        
        # 隨機產生E_uug
        tmp_E_uug[0][hour_i] = random.uniform(0, max(D_user[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_uus[0][hour_i], 0))
        
        # E_ubo必須用來滿足D_user
        tmp_E_ubo[0][hour_i] = max(D_user[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_uus[0][hour_i] - tmp_E_uug[0][hour_i], 0)
        tmp_s_on_t[0][hour_i+1] = tmp_s_on_t[0][hour_i] + tmp_E_sfg[0][hour_i] + tmp_E_sfo[0][hour_i] - tmp_E_uus[0][hour_i] - tmp_E_sso[0][hour_i] - tmp_E_vfs[0][hour_i] + ParaOfRenergy[DAY][hour_i]
        tmp_ev_on_t[0][hour_i+1] = tmp_ev_on_t[0][hour_i] + tmp_E_vfg[0][hour_i] + tmp_E_vfo[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_vso[0][hour_i]
        
        # 將產生候選解加到人口中
        population[i].E_uug.append(tmp_E_uug[0][hour_i])
        population[i].E_sfg.append(tmp_E_sfg[0][hour_i])
        population[i].E_uus.append(tmp_E_uus[0][hour_i])
        population[i].E_sso.append(tmp_E_sso[0][hour_i])
        population[i].E_sfo.append(tmp_E_sfo[0][hour_i])
        population[i].E_ubo.append(tmp_E_ubo[0][hour_i])
        population[i].E_vfg.append(tmp_E_vfg[0][hour_i])
        population[i].E_uuv.append(tmp_E_uuv[0][hour_i])
        population[i].E_vfs.append(tmp_E_vfs[0][hour_i])
        population[i].E_vso.append(tmp_E_vso[0][hour_i])
        population[i].E_vfo.append(tmp_E_vfo[0][hour_i])
        population[i].theta_s.append(tmp_theta_s[0][hour_i])
        population[i].theta_ev.append(tmp_theta_ev[0][hour_i])
        population[i].check_s_on_t.append( tmp_s_on_t[0][hour_i +1])
        population[i].check_ev_on_t.append( tmp_ev_on_t[0][hour_i +1])
    
# Step1.5:檢查是否有不可行解    
for i in range( NumOfPerson):
    ChechLegalOrNot( population[i])
    
# Step2:計算適應度值    
for i in range( NumOfPerson):
    population[i].fitness = calTheFit( population[i])
    

# Step2.5:存取t_0世代下最好的解

FitnessArray = list()                                   # 用來存當世代所有候選解的fitness
for i in range( NumOfPerson):
    FitnessArray.append(population[i].fitness)          # temp矩陣存當世代所有人的fitness
    
IndexOfBestCS = FitnessArray.index( min(FitnessArray))              # 找到第0世代最好的解 並返回索引值

Hist_BestFit.append( population[IndexOfBestCS].fitness)
Hist_AvgFit.append( sum( FitnessArray)/len( FitnessArray))


BestPerson = list()
BestPerson.append( population[IndexOfBestCS])

list_Par_a = list( )

time_start_FLAG = time.time()

# Step3:主迴圈開始    
for iteration in range( MaxOfIteration):
    
    OffSpring = list()                                                  # 宣告子代
    
    time_iter_start = time.time()
    
    ### Step4:更新參數

    if iteration < ( 0.5 * 10000):
        Par_a = (iteration / ( 10000 + 0.0001)) * 2
    else:
        if Par_a > 0:
            Par_a = (2 - ( (2 * iteration) / ( 10000 + 0.0001))) * 2
        else:
            Par_a = 0
    
        
    
    Par_A = 2 * Par_a * np.random.random() - Par_a
    Par_C = 2 * np.random.random()
    Par_l = np.random.uniform( -1, 1)
    Par_p = np.random.random()
    
    
    for i in range( NumOfPerson):                                       # 1次產生1條子代
        OffSpring.append(person())                      
            
        # 初始化儲電設備與電動車每小時的電量 
        tmp_E_uug, tmp_E_sfg, tmp_E_uus = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_E_sso, tmp_E_sfo, tmp_E_ubo = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_E_vfg, tmp_E_uuv, tmp_E_vfs = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_E_vso, tmp_E_vfo, tmp_theta_s, tmp_theta_ev = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_s_on_t, tmp_ev_on_t = np.zeros((1, Hours + 1)), np.zeros((1, Hours + 1))
    
        
        if DAY == 0:
            tmp_s_on_t[0][0] = E_s_min
            tmp_ev_on_t[0][0] = E_ev_min
            
        else:
#            file = load_workbook( PathOfFile + 'Remain0416.xlsx')
#            sheet_ranges = file['HowManyStillHave_WOA']
#            
#            for k in range( len( DAY_List)):
#                if DAY == list(enumerate(DAY_List))[k][0]:
#                    tmp_s_on_t[0][0] = sheet_ranges.cell( row = k + 1, column = 2).value
#                    tmp_ev_on_t[0][0] = sheet_ranges.cell( row = k + 1, column = 3).value
#                    break
            tmp_s_on_t[0][0] = excel_s_on_t    
            tmp_ev_on_t[0][0] = excel_ev_on_t
                
        OffSpring[0].check_s_on_t.append( tmp_s_on_t[0][0])
        OffSpring[0].check_ev_on_t.append( tmp_ev_on_t[0][0])
        
        # Step5:找鄰近解
        if Par_p < 0.5:
            
            # 過去最佳(2.1)
            if abs( Par_A) < 1:
            
                for hour_i in range( Hours):
                    
                    # theta_s
                    tmp_theta_s[0][hour_i] = BestPerson[0].theta_s[hour_i]
                    # theta_ev
                    if hour_i > 6 and hour_i < 17:                          # 上班時段
                        tmp_theta_ev[0][hour_i] = 0
                    else:                                                   # 非上班時段
                        tmp_theta_ev[0][hour_i] = BestPerson[0].theta_ev[hour_i]
                    
                    tmp = BestPerson[0].E_uug[hour_i]
                    tmp_E_uug[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_uug[hour_i]), 0)
                    tmp = BestPerson[0].E_sfg[hour_i]
                    tmp_E_sfg[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_sfg[hour_i]), 0)
                    tmp = BestPerson[0].E_uus[hour_i]
                    tmp_E_uus[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_uus[hour_i]), 0)
                    tmp = BestPerson[0].E_sso[hour_i]
                    tmp_E_sso[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_sso[hour_i]), 0)                                        
                    tmp = BestPerson[0].E_sfo[hour_i]
                    tmp_E_sfo[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_sfo[hour_i]), 0)
                    tmp = BestPerson[0].E_ubo[hour_i]
                    tmp_E_ubo[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_ubo[hour_i]), 0)
                    tmp = BestPerson[0].E_vfg[hour_i]
                    tmp_E_vfg[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_vfg[hour_i]), 0)
                    tmp = BestPerson[0].E_uuv[hour_i]
                    tmp_E_uuv[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_uuv[hour_i]), 0)
                    tmp = BestPerson[0].E_vfs[hour_i]
                    tmp_E_vfs[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_vfs[hour_i]), 0)
                    tmp = BestPerson[0].E_vso[hour_i]
                    tmp_E_vso[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_vso[hour_i]), 0)
                    tmp = BestPerson[0].E_vfo[hour_i]
                    tmp_E_vfo[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_vfo[hour_i]), 0)
            
            # 過去隨機解(2.8)
            else:
                
                # 挑一個隨機候選解
                IndexOfRandCS = np.random.randint(0, NumOfPerson)
                
                for hour_i in range( Hours):
                    
                    # theta_s
#                    tmp_theta_s[0][hour_i] = np.random.randint(0, 2)
                    tmp_theta_s[0][hour_i] = population[i].theta_s[hour_i]
                    # theta_ev
                    if hour_i > 6 and hour_i < 17:                          # 上班時段
                        tmp_theta_ev[0][hour_i] = 0
                    else:                                                   # 非上班時段
#                        tmp_theta_ev[0][hour_i] = np.random.randint(0, 2)
                        tmp_theta_ev[0][hour_i] = population[i].theta_ev[hour_i]
                        
                    tmp = population[IndexOfRandCS].E_uug[hour_i]
                    tmp_E_uug[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_uug[hour_i]), 0)
                    tmp = population[IndexOfRandCS].E_sfg[hour_i]
                    tmp_E_sfg[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_sfg[hour_i]), 0)
                    tmp = population[IndexOfRandCS].E_uus[hour_i]
                    tmp_E_uus[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_uus[hour_i]), 0)
                    tmp = population[IndexOfRandCS].E_sso[hour_i]
                    tmp_E_sso[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_sso[hour_i]), 0)                                        
                    tmp = population[IndexOfRandCS].E_sfo[hour_i]
                    tmp_E_sfo[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_sfo[hour_i]), 0)
                    tmp = population[IndexOfRandCS].E_ubo[hour_i]
                    tmp_E_ubo[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_ubo[hour_i]), 0)
                    tmp = population[IndexOfRandCS].E_vfg[hour_i]
                    tmp_E_vfg[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_vfg[hour_i]), 0)
                    tmp = population[IndexOfRandCS].E_uuv[hour_i]
                    tmp_E_uuv[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_uuv[hour_i]), 0)
                    tmp = population[IndexOfRandCS].E_vfs[hour_i]
                    tmp_E_vfs[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_vfs[hour_i]), 0)
                    tmp = population[IndexOfRandCS].E_vso[hour_i]
                    tmp_E_vso[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_vso[hour_i]), 0)
                    tmp = population[IndexOfRandCS].E_vfo[hour_i]
                    tmp_E_vfo[0][hour_i] = max(tmp - Par_A * abs( Par_C * tmp - population[i].E_vfo[hour_i]), 0)
                    
        # 圓心最佳解
        else:
        
            for hour_i in range( Hours):
                
                # theta_s
                tmp_theta_s[0][hour_i] = BestPerson[0].theta_s[hour_i]
                # theta_ev
                if hour_i > 6 and hour_i < 17:                          # 上班時段
                    tmp_theta_ev[0][hour_i] = 0
                else:                                                   # 非上班時段
                    tmp_theta_ev[0][hour_i] = BestPerson[0].theta_ev[hour_i]
                
                
                tmp_E_uug[0][hour_i] = CycleMove( BestPerson[0].E_uug[hour_i], population[i].E_uug[hour_i])
                tmp_E_sfg[0][hour_i] = CycleMove( BestPerson[0].E_sfg[hour_i], population[i].E_sfg[hour_i])
                tmp_E_uus[0][hour_i] = CycleMove( BestPerson[0].E_uus[hour_i], population[i].E_uus[hour_i])
                tmp_E_sso[0][hour_i] = CycleMove( BestPerson[0].E_sso[hour_i], population[i].E_sso[hour_i])
                tmp_E_sfo[0][hour_i] = CycleMove( BestPerson[0].E_sfo[hour_i], population[i].E_sfo[hour_i])
                tmp_E_ubo[0][hour_i] = CycleMove( BestPerson[0].E_ubo[hour_i], population[i].E_ubo[hour_i])
                tmp_E_vfg[0][hour_i] = CycleMove( BestPerson[0].E_vfg[hour_i], population[i].E_vfg[hour_i])
                tmp_E_ubo[0][hour_i] = CycleMove( BestPerson[0].E_ubo[hour_i], population[i].E_ubo[hour_i])
                tmp_E_uuv[0][hour_i] = CycleMove( BestPerson[0].E_uuv[hour_i], population[i].E_uuv[hour_i])
                tmp_E_vfs[0][hour_i] = CycleMove( BestPerson[0].E_vfs[hour_i], population[i].E_vfs[hour_i])
                tmp_E_vso[0][hour_i] = CycleMove( BestPerson[0].E_vso[hour_i], population[i].E_vso[hour_i])
                tmp_E_vfo[0][hour_i] = CycleMove( BestPerson[0].E_vfo[hour_i], population[i].E_vfo[hour_i])
                
              
        for hour_i in range( Hours):
            OffSpring[i].E_uug.append(tmp_E_uug[0][hour_i])
            OffSpring[i].E_sfg.append(tmp_E_sfg[0][hour_i])
            OffSpring[i].E_uus.append(tmp_E_uus[0][hour_i])
            OffSpring[i].E_sso.append(tmp_E_sso[0][hour_i])
            OffSpring[i].E_sfo.append(tmp_E_sfo[0][hour_i])
            OffSpring[i].E_ubo.append(tmp_E_ubo[0][hour_i])
            OffSpring[i].E_vfg.append(tmp_E_vfg[0][hour_i])
            OffSpring[i].E_uuv.append(tmp_E_uuv[0][hour_i])
            OffSpring[i].E_vfs.append(tmp_E_vfs[0][hour_i])
            OffSpring[i].E_vso.append(tmp_E_vso[0][hour_i])
            OffSpring[i].E_vfo.append(tmp_E_vfo[0][hour_i])
            OffSpring[i].theta_s.append(tmp_theta_s[0][hour_i])
            OffSpring[i].theta_ev.append(tmp_theta_ev[0][hour_i])
        
    
    ##### Step5.5:嘗試 IWoa 多樣性變異操作 ##########
    IndexOfRandCS = np.random.randint(0, NumOfPerson)
    
    
    # 產生人口前先初始化 
    tmp_E_uug, tmp_E_sfg, tmp_E_uus = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_E_sso, tmp_E_sfo, tmp_E_ubo = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_E_vfg, tmp_E_uuv, tmp_E_vfs = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_E_vso, tmp_E_vfo, tmp_theta_s, tmp_theta_ev = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_s_on_t, tmp_ev_on_t = np.zeros((1, Hours + 1)), np.zeros((1, Hours + 1))
    
    if DAY == 0:
        tmp_s_on_t[0][0] = E_s_min
        tmp_ev_on_t[0][0] = E_ev_min
    else:
#        file = load_workbook( PathOfFile + 'Remain0416.xlsx')
#        sheet_ranges = file['HowManyStillHave_WOA']
#    
#        for k in range( len( DAY_List)):
#            if DAY == list(enumerate(DAY_List))[k][0]:
#                tmp_s_on_t[0][0] = (sheet_ranges.cell( row = k + 1, column = 2).value)
#                tmp_ev_on_t[0][0] = (sheet_ranges.cell( row = k + 1, column = 3).value)
#                break
            
        tmp_s_on_t[0][0] = excel_s_on_t    
        tmp_ev_on_t[0][0] = excel_ev_on_t
        
        
            
    OffSpring[IndexOfRandCS].check_s_on_t.append( tmp_s_on_t[0][0])
    OffSpring[IndexOfRandCS].check_ev_on_t.append( tmp_ev_on_t[0][0])
    
    for hour_i in range( Hours):
        
        # 確認儲電設備
        if tmp_s_on_t[0][hour_i] <= E_s_min:                  # 儲電設備沒存量不能放電
            tmp_theta_s[0][hour_i] = 1
        elif tmp_s_on_t[0][hour_i] + ParaOfRenergy[DAY][hour_i] >= E_s_max:   # 儲電設備沒容量不能儲電
            tmp_theta_s[0][hour_i] = 0
        else:
            tmp_theta_s[0][hour_i] = random.randint(0, 1)
        
        # 確認電動
        if tmp_ev_on_t[0][hour_i] <= E_ev_min:
            tmp_theta_ev[0][hour_i] = 1
        elif tmp_ev_on_t[0][hour_i] >= E_ev_max or Theta_Work[DAY][hour_i] == 1:
            tmp_theta_ev[0][hour_i] = 0
        else:
            tmp_theta_ev[0][hour_i] = random.randint(0, 1)
        
        # 產生儲放電值
        
        # 當儲電設備儲電，電動車儲電
        if tmp_theta_s[0][hour_i] == 1 and tmp_theta_ev[0][hour_i] == 1:            
            
            # 儲電設備
            tmp_E_sfg[0][hour_i] = random.uniform(0, min( (E_s_max - tmp_s_on_t[0][hour_i] - ParaOfRenergy[DAY][hour_i]), Rate_s))
            tmp_E_sfo[0][hour_i] = random.uniform(0, min( max(E_s_max - tmp_s_on_t[0][hour_i] - tmp_E_sfg[0][hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s))        
            
            #電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_vfg[0][hour_i] = random.uniform(0, min(E_ev_max - tmp_ev_on_t[0][hour_i], Rate_ev))
                tmp_E_vfo[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfg[0][hour_i], 0), Rate_ev))
        
        # 當儲電設備放電，電動車儲電
        elif tmp_theta_s[0][hour_i] == 0 and tmp_theta_ev[0][hour_i] == 1:
            
            # 儲電設備
            tmp_E_uus[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min( tmp_s_on_t[0][hour_i], Rate_s)))
            tmp_E_sso[0][hour_i] = random.uniform(0, min(max( tmp_s_on_t[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s))
            
            #電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_vfs[0][hour_i] = random.uniform(0, min(E_ev_max - tmp_ev_on_t[0][hour_i], min(max( tmp_s_on_t[0][hour_i] - tmp_E_sso[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s)))
                tmp_E_vfg[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfs[0][hour_i],0), Rate_ev))
                tmp_E_vfo[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfg[0][hour_i] - tmp_E_vfs[0][hour_i], 0), Rate_ev))
        
        # 當儲電設備儲電，電動車放電
        elif tmp_theta_s[0][hour_i] == 1 and tmp_theta_ev[0][hour_i] == 0:
        
            # 儲電設備
            tmp_E_sfg[0][hour_i] = random.uniform(0, min( (E_s_max - tmp_s_on_t[0][hour_i] - ParaOfRenergy[DAY][hour_i]), Rate_s))
            tmp_E_sfo[0][hour_i] = random.uniform(0, min( max(E_s_max - tmp_s_on_t[0][hour_i] - tmp_E_sfg[0][hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s))
            
            #儲電設備
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_uuv[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min(tmp_ev_on_t[0][hour_i] - E_ev_min, Rate_ev)))
                tmp_E_vso[0][hour_i] = random.uniform(0, min(max(tmp_ev_on_t[0][hour_i] - tmp_E_uuv[0][hour_i] - E_ev_min, 0), Rate_ev))
        
        # 當儲電設備放電，電動車放電
        elif tmp_theta_s[0][hour_i] == 0 and tmp_theta_ev[0][hour_i] == 0:
        
            # 儲電設備
            tmp_E_uus[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min( tmp_s_on_t[0][hour_i], Rate_s)))
            tmp_E_sso[0][hour_i] = random.uniform(0, min(max( tmp_s_on_t[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s))
            # 電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_uuv[0][hour_i] = random.uniform(0, min(max(0, D_user[0][hour_i] - tmp_E_uus[0][hour_i]), min(tmp_ev_on_t[0][hour_i] - E_ev_min, Rate_ev)))
                tmp_E_vso[0][hour_i] = random.uniform(0, min(max(tmp_ev_on_t[0][hour_i] - tmp_E_uuv[0][hour_i] - E_ev_min, 0), Rate_ev))
        else:
            print('Error in initial!!')
        
        # 隨機產生E_uug
        tmp_E_uug[0][hour_i] = random.uniform(0, max(D_user[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_uus[0][hour_i], 0))
        
        # E_ubo必須用來滿足D_user
        tmp_E_ubo[0][hour_i] = max(D_user[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_uus[0][hour_i] - tmp_E_uug[0][hour_i], 0)
        tmp_s_on_t[0][hour_i+1] = tmp_s_on_t[0][hour_i] + tmp_E_sfg[0][hour_i] + tmp_E_sfo[0][hour_i] - tmp_E_uus[0][hour_i] - tmp_E_sso[0][hour_i] - tmp_E_vfs[0][hour_i] + ParaOfRenergy[DAY][hour_i]
        tmp_ev_on_t[0][hour_i+1] = tmp_ev_on_t[0][hour_i] + tmp_E_vfg[0][hour_i] + tmp_E_vfo[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_vso[0][hour_i]
        
        # 將產生候選解加到人口中
        OffSpring[IndexOfRandCS].E_uug.append(tmp_E_uug[0][hour_i])
        OffSpring[IndexOfRandCS].E_sfg.append(tmp_E_sfg[0][hour_i])
        OffSpring[IndexOfRandCS].E_uus.append(tmp_E_uus[0][hour_i])
        OffSpring[IndexOfRandCS].E_sso.append(tmp_E_sso[0][hour_i])
        OffSpring[IndexOfRandCS].E_sfo.append(tmp_E_sfo[0][hour_i])
        OffSpring[IndexOfRandCS].E_ubo.append(tmp_E_ubo[0][hour_i])
        OffSpring[IndexOfRandCS].E_vfg.append(tmp_E_vfg[0][hour_i])
        OffSpring[IndexOfRandCS].E_uuv.append(tmp_E_uuv[0][hour_i])
        OffSpring[IndexOfRandCS].E_vfs.append(tmp_E_vfs[0][hour_i])
        OffSpring[IndexOfRandCS].E_vso.append(tmp_E_vso[0][hour_i])
        OffSpring[IndexOfRandCS].E_vfo.append(tmp_E_vfo[0][hour_i])
        OffSpring[IndexOfRandCS].theta_s.append(tmp_theta_s[0][hour_i])
        OffSpring[IndexOfRandCS].theta_ev.append(tmp_theta_ev[0][hour_i])
        OffSpring[IndexOfRandCS].check_s_on_t.append( tmp_s_on_t[0][hour_i +1])
        OffSpring[IndexOfRandCS].check_ev_on_t.append( tmp_ev_on_t[0][hour_i +1])
    
    #####
    
        
    # Step6:修正不可行解
    for i in range( NumOfPerson):
    
        OffSpring[i].check_s_on_t = list()
        OffSpring[i].check_ev_on_t = list()
        
        if DAY == 0:
            OffSpring[i].check_s_on_t.append(E_s_min)
            OffSpring[i].check_ev_on_t.append(E_ev_min) 
        else:
            file = load_workbook( PathOfFile + 'Remain0416.xlsx')
            sheet_ranges = file['HowManyStillHave_WOA']
            
            for k in range( len( DAY_List)):
                if DAY == list(enumerate(DAY_List))[k][0]:
                    OffSpring[i].check_s_on_t.append( sheet_ranges.cell( row = k + 1, column = 2).value)
                    OffSpring[i].check_ev_on_t.append( sheet_ranges.cell( row = k + 1, column = 3).value) 
                    break
                        
        for hour_i in range( Hours):
            OffSpring[i].check_s_on_t.append( 0)
            OffSpring[i].check_ev_on_t.append( 0)
        
        for hour_i in range( Hours):
            
            # 1. 檢查儲放電設備是否有放電錯誤
            if OffSpring[i].theta_s[hour_i] == 0:
                if OffSpring[i].E_sfg[hour_i] > 0:
                    OffSpring[i].E_sfg[hour_i] = 0

                if OffSpring[i].E_sfo[hour_i] > 0:
                    OffSpring[i].E_sfo[hour_i] = 0
            else:
                if OffSpring[i].E_uus[hour_i] > 0:
                    OffSpring[i].E_uug[hour_i] = OffSpring[i].E_uug[hour_i] + OffSpring[i].E_uus[hour_i]
                    OffSpring[i].E_uus[hour_i] = 0
                
                if OffSpring[i].E_sso[hour_i] > 0:
                    OffSpring[i].E_sso[hour_i] = 0
                
                if OffSpring[i].E_vfs[hour_i] > 0:
                    OffSpring[i].E_vfg[hour_i] = OffSpring[i].E_vfg[hour_i] + OffSpring[i].E_vfs[hour_i]
                    OffSpring[i].E_vfs[hour_i] = 0
            
            # 2. 檢查電動車是否由儲放電錯誤
            # 如果在上班時段
            if Theta_Work[DAY][hour_i] == 1:
                if OffSpring[i].E_vfg[hour_i] > 0:
                    OffSpring[i].E_vfg[hour_i] = 0
                if OffSpring[i].E_vfs[hour_i] > 0:
                    OffSpring[i].E_sso[hour_i] = OffSpring[i].E_sso[hour_i] + OffSpring[i].E_vfs[hour_i]
                    OffSpring[i].E_vfs[hour_i] = 0
                if OffSpring[i].E_vfo[hour_i] > 0:
                    OffSpring[i].E_vfo[hour_i] = 0
                if OffSpring[i].E_uuv[hour_i] > 0:
                    OffSpring[i].E_uug[hour_i] = OffSpring[i].E_uuv[hour_i] + OffSpring[i].E_uuv[hour_i]
                    OffSpring[i].E_uuv[hour_i] = 0
                if OffSpring[i].E_vso[hour_i] > 0:
                    OffSpring[i].E_vso[hour_i] = 0
     
            if OffSpring[i].theta_ev[hour_i] == 0:
                if OffSpring[i].E_vfg[hour_i] > 0:
                    OffSpring[i].E_vfg[hour_i] = 0
                if OffSpring[i].E_vfs[hour_i] > 0:
                    OffSpring[i].E_sso[hour_i] = OffSpring[i].E_sso[hour_i] + OffSpring[i].E_vfs[hour_i]
                    OffSpring[i].E_vfs[hour_i] = 0
                if OffSpring[i].E_vfo[hour_i] > 0:
                    OffSpring[i].E_vfo[hour_i] = 0
            else:
                if OffSpring[i].E_uuv[hour_i] > 0:
                    OffSpring[i].E_uug[hour_i] = OffSpring[i].E_uug[hour_i] + OffSpring[i].E_uuv[hour_i]
                    OffSpring[i].E_uuv[hour_i] = 0
                if OffSpring[i].E_vso[hour_i] > 0:
                    OffSpring[i].E_vso[hour_i] = 0
            
            # 3. 檢查儲電設備與電動車容量限制
            tmp_s = OffSpring[i].check_s_on_t[hour_i] + OffSpring[i].E_sfg[hour_i] + OffSpring[i].E_sfo[hour_i] - OffSpring[i].E_uus[hour_i] - OffSpring[i].E_sso[hour_i] - OffSpring[i].E_vfs[hour_i] + ParaOfRenergy[DAY][hour_i]
            tmp_ev = OffSpring[i].check_ev_on_t[hour_i] + OffSpring[i].E_vfg[hour_i] + OffSpring[i].E_vfs[hour_i] + OffSpring[i].E_vfo[hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_vso[hour_i]
            

            
            # 當儲電設備發生沒電還放電或是滿載還儲電的話, 就要修正
            if ((OffSpring[i].check_s_on_t[ hour_i + 1] < E_s_min) or (OffSpring[i].check_s_on_t[ hour_i + 1] > E_s_max)) or ((OffSpring[i].check_ev_on_t[ hour_i + 1] < E_ev_min) or ( OffSpring[i].check_ev_on_t[ hour_i + 1] > E_ev_max)):
                                         
                OffSpring[i].E_uug[hour_i] = 0
                OffSpring[i].E_sfg[hour_i] = 0
                OffSpring[i].E_uus[hour_i] = 0
                OffSpring[i].E_sso[hour_i] = 0
                OffSpring[i].E_sfo[hour_i] = 0
                OffSpring[i].E_ubo[hour_i] = 0
                OffSpring[i].E_vfg[hour_i] = 0
                OffSpring[i].E_uuv[hour_i] = 0
                OffSpring[i].E_vfs[hour_i] = 0
                OffSpring[i].E_vso[hour_i] = 0
                OffSpring[i].E_vfo[hour_i] = 0
                
                # 儲電設備沒存量不能放電
                if OffSpring[i].check_s_on_t[ hour_i] <= E_s_min:
                    OffSpring[i].theta_s[hour_i] = 1
                # 儲電設備沒容量不能儲電
                elif OffSpring[i].check_s_on_t[ hour_i] + ParaOfRenergy[DAY][hour_i] >= E_s_max:
                    OffSpring[i].theta_s[hour_i] = 0
                else:
                    OffSpring[i].theta_s[hour_i] = np.random.randint(0, 2)
            
                # 電動車沒存量不能放電
                if OffSpring[i].check_ev_on_t[ hour_i] <= E_ev_min:
                    OffSpring[i].theta_ev[hour_i] = 1
                # 電動車沒容量不能儲電
                elif OffSpring[i].check_ev_on_t[ hour_i] >= E_ev_max:
                    OffSpring[i].theta_ev[hour_i] = 0
                else:
                    OffSpring[i].theta_ev[hour_i] = np.random.randint(0, 2)
                
                # 儲電設備及電動車同時儲電情況
                if OffSpring[i].theta_s[hour_i] == 1 and OffSpring[i].theta_ev[hour_i] == 1:
                    # 儲電設備
                    OffSpring[i].E_sfg[hour_i] = random.uniform( 0, min( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - ParaOfRenergy[DAY][hour_i], Rate_s) )
                    OffSpring[i].E_sfo[hour_i] = random.uniform( 0, min( max( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_sfg[hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s))
                    # 電動車
                    if Theta_Work[DAY][hour_i] == 0:
                        OffSpring[i].E_vfg[hour_i] = random.uniform( 0, min( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i], Rate_ev))
                        OffSpring[i].E_vfo[hour_i] = random.uniform( 0, min( max( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_vfg[hour_i], 0), Rate_ev))
                
                # 產生儲放電值，當儲電設備放電，電動車儲電
                elif OffSpring[i].theta_s[hour_i] == 0 and OffSpring[i].theta_ev[hour_i] == 1:
                    # 儲電設備
                    OffSpring[i].E_uus[hour_i] = random.uniform( 0, min( D_user[0][hour_i], min( OffSpring[i].check_s_on_t[ hour_i], Rate_s)))
                    OffSpring[i].E_sso[hour_i] = random.uniform( 0, min( max( OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_uus[hour_i], 0), Rate_s))
                    # 電動車
                    if Theta_Work[DAY][hour_i] == 0:
                        OffSpring[i].E_vfs[hour_i] = random.uniform( 0, min( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i], min( max(  OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_sso[hour_i] - OffSpring[i].E_uus[hour_i],0), Rate_s)))
                        OffSpring[i].E_vfg[hour_i] = random.uniform( 0, min( max( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_vfs[hour_i], 0), Rate_ev))
                        OffSpring[i].E_vfo[hour_i] = random.uniform( 0, min( max( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_vfg[hour_i] - OffSpring[i].E_vfs[hour_i], 0), Rate_ev))
                
                # 產生儲放電值，當儲電設備儲電，電動車放電
                elif OffSpring[i].theta_s[hour_i] == 1 and OffSpring[i].theta_ev[hour_i] == 0:
                    # 儲電設備
                    OffSpring[i].E_sfg[hour_i] = random.uniform(0, min( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - ParaOfRenergy[DAY][hour_i], Rate_s))
                    OffSpring[i].E_sfo[hour_i] = random.uniform( 0, min( max( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_sfg[hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s ))
                    # 電動車
                    if Theta_Work[DAY][hour_i] == 0:
                        OffSpring[i].E_uuv[hour_i] = random.uniform(0, min( D_user[0][hour_i], min( OffSpring[i].check_ev_on_t[ hour_i] - E_ev_min, Rate_ev)))
                        OffSpring[i].E_vso[hour_i] = random.uniform(0, min(max( OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_uuv[hour_i] - E_ev_min, 0), Rate_ev))
                
                # 產生儲放電值，當儲電設備放電，電動車放電
                else:
                    # 儲電設備
                    OffSpring[i].E_uus[hour_i] = random.uniform( 0, min( D_user[0][hour_i], min( OffSpring[i].check_s_on_t[ hour_i], Rate_s)))
                    OffSpring[i].E_sso[hour_i] = random.uniform( 0, min( max( OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_uus[hour_i], 0), Rate_s))
                    # 電動車
                    if Theta_Work[DAY][hour_i] == 0:
                        OffSpring[i].E_uuv[hour_i] = random.uniform( 0, min( max( 0, D_user[0][hour_i] - OffSpring[i].E_uus[hour_i]), min( OffSpring[i].check_ev_on_t[ hour_i] - E_ev_min, Rate_ev)))
                        OffSpring[i].E_vso[hour_i] = random.uniform( 0, min( max( OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_uuv[hour_i] - E_ev_min, 0), Rate_ev))
                
                # 隨機產生E_uug
                OffSpring[i].E_uug[hour_i] = random.uniform( 0, max( D_user[0][hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_uus[hour_i], 0))
                # E_ubo必須用來滿足D_user
                OffSpring[i].E_ubo[hour_i] = max( D_user[0][hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_uus[hour_i] - OffSpring[i].E_uug[hour_i], 0)
                    
            # 更新結果
            tmp_s = OffSpring[i].check_s_on_t[hour_i] + OffSpring[i].E_sfg[hour_i] + OffSpring[i].E_sfo[hour_i] - OffSpring[i].E_uus[hour_i] - OffSpring[i].E_sso[hour_i] - OffSpring[i].E_vfs[hour_i] + ParaOfRenergy[DAY][hour_i]
            tmp_ev = OffSpring[i].check_ev_on_t[hour_i] + OffSpring[i].E_vfg[hour_i] + OffSpring[i].E_vfs[hour_i] + OffSpring[i].E_vfo[hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_vso[hour_i]
                        
            OffSpring[i].check_s_on_t[hour_i +1] = tmp_s
            OffSpring[i].check_ev_on_t[hour_i +1] = tmp_ev
        
    # Step7:檢查修正後是否有有不可行解        
    for i in range( NumOfPerson):
        ChechLegalOrNot( OffSpring[i])
     
    # Step8:計算適應度值    
    for i in range( NumOfPerson):
        OffSpring[i].fitness = calTheFit( OffSpring[i])  
        
    # Step9:確認有無更新最佳解
    
    FitnessArray = list()                                   # 用來存當世代所有候選解的fitness
    for i in range( NumOfPerson):
        FitnessArray.append(OffSpring[i].fitness)          # temp矩陣存當世代所有人的fitness
    
    # 若t世代下最佳解有比歷史最佳解好, 就取代歷史最佳解 
    if min(FitnessArray) <= Hist_BestFit[-1]:
        
        IndexOfBestCS = FitnessArray.index( min(FitnessArray))              # 找到第0世代最好的解 並返回索引值
        Hist_BestFit.append( OffSpring[IndexOfBestCS].fitness)
    
        BestPerson = list()
        BestPerson.append( OffSpring[IndexOfBestCS])

    else:
        Hist_BestFit.append( Hist_BestFit[-1])
    
    
    # 當代下的平均解照樣存取
    Hist_AvgFit.append( sum( FitnessArray)/len( FitnessArray))
    time_iter_end = time.time()
    time_end_FLAG = time.time()
    
    if iteration % 100 == 0:
        print( 'now in', iteration, 'iteration')
        print( '100 iteration spent:', time_end_FLAG - time_start_FLAG )
        time_start_FLAG = time.time()
#        print( "now are spent :", round(time_FLAG - time_start, 4))
    
#    if (( time_FLAG - time_start) >= 825):
#        break
    

# ==== 演算法結束 ====
# Step8:看最後的解
BestCSSolution = np.zeros(( 13, Hours))
for i in range( Hours):
    BestCSSolution[0][i] = BestPerson[0].E_uug[i]
    BestCSSolution[1][i] = BestPerson[0].E_sfg[i]
    BestCSSolution[2][i] = BestPerson[0].E_uus[i]
    BestCSSolution[3][i] = BestPerson[0].E_sso[i]
    BestCSSolution[4][i] = BestPerson[0].E_sfo[i]
    BestCSSolution[5][i] = BestPerson[0].E_ubo[i]
    BestCSSolution[6][i] = BestPerson[0].E_vfg[i]
    BestCSSolution[7][i] = BestPerson[0].E_uuv[i]
    BestCSSolution[8][i] = BestPerson[0].E_vfs[i]
    BestCSSolution[9][i] = BestPerson[0].E_vso[i]
    BestCSSolution[10][i] = BestPerson[0].E_vfo[i]
    BestCSSolution[11][i] = BestPerson[0].theta_s[i]
    BestCSSolution[12][i] = BestPerson[0].theta_ev[i]

# 將儲電設備與電動車最後的電量儲存起來 
ToTxt_s = BestPerson[0].check_s_on_t[-1]
ToTxt_ev = BestPerson[0].check_ev_on_t[-1]

file = load_workbook( PathOfFile + 'Remain0416.xlsx')
sheet_ranges = file['HowManyStillHave_WOA']

for k in range( len( DAY_List)):
    if DAY == list(enumerate(DAY_List))[k][0]:
        sheet_ranges.cell(row = k + 2, column = 2).value = ToTxt_s
        sheet_ranges.cell(row = k + 2, column = 3).value = ToTxt_ev
        break


file.save( 'Remain0416.xlsx')
file.close()


time_end = time.time()
print( "time:", time_end - time_start)


