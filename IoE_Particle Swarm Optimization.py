"""
Created on MON 0204 22:04:46 2019
@author: ChihChi
"""
import pandas as pd
import numpy as np
import random
import math
import time
from openpyxl import Workbook
from openpyxl import load_workbook

List_bestValue = list()
List_time = list()



# 參數設定
NumOfPerson = 20
Hours = 24
E_s_min = 0
E_ev_min = 12.5
E_s_max = 13.5
E_ev_max = 50
Rate_s = 5
Rate_ev = 10 
Beta_s = 0.1
Beta_ev = 0.1

# 目前週幾
# DAY = 0 -> Monday, DAY = 1 -> Tuesday, DAY = 2 -> Wed
DAY = 3
DAY_List = [ 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']

Hist_BestFit = list()           # 當代最佳
Hist_AvgFit = list()            # 當代平均
Hist_PreBest = list()           # 過去最佳


# 演算法參數
MaxOfIteration = 25000

Par_w = 0.8
Par_c1 = 2
Par_c2 = 2
Par_Vmax = 2                # 鳥群的速度上限
Par_SetRange = 0.3          # 初始的速度上下界
#    Rate_s = 5
#    Rate_ev = 8

#a = random.randint( 1, 15000000)
a = 106076

# 設定亂數種子
random.seed( a)
np.random.seed( a)



# 建立物件
class person:
    def __init__(self):
        self.E_uug = list()
        self.E_sfg = list()
        self.E_uus = list()
        self.E_sso = list()
        self.E_sfo = list()
        self.E_ubo = list()
        self.E_vfg = list()
        self.E_uuv = list()
        self.E_vfs = list()
        self.E_vso = list()
        self.E_vfo = list()
        self.theta_s = list()
        self.theta_ev = list()
        self.check_s_on_t = list()
        self.check_ev_on_t = list()
        self.fitness = 0
        self.velocity = np.zeros(( 13, Hours))
        self.matrix = np.zeros(( 13, Hours))
        self.previousBestMatrix = np.zeros(( 13, Hours))
        self.previousFitness = 0
        
        
# =============================================================================        
PathOfFile = 'C:/Users/juice/Desktop/MyProposal/IoE_Program/'

# == 可再生能源 ==
target = pd.read_excel( PathOfFile + 'RenewableEnergy.xlsx', dtype = 'float64')
ParaOfRenergy = target.values

# == 實時電價 ==
target = pd.read_excel( PathOfFile + 'RTP.xlsx', dtype = 'float64')
ParaOfRTP = target.values

# == 市場價格 ==
target = pd.read_excel( PathOfFile + 'MarketingPrice.xlsx', dtype = 'float64')
ParaOfMarketP = target.values

# == D_user ==
target = pd.read_excel( PathOfFile + 'Duser.xlsx', dtype = 'float64')
D_user = target.values

# == Theta_Work ==
target = pd.read_excel( PathOfFile + 'Theta_work.xlsx', dtype = 'float64')
Theta_Work = target.values

# =============================================================================

# function
def calTheFit( person):
    
    C_grid, C_user, R_profit, R_capital = 0, 0, 0, 0
    for hour_i in range( Hours):
        C_grid = C_grid + ParaOfRTP[DAY][hour_i] * ( person.E_uug[hour_i] + ( person.E_sfg[hour_i] / ( 1 - Beta_s)) + ( person.E_vfg[hour_i] / (1 - Beta_ev)))
        C_user = C_user + ParaOfMarketP[DAY][hour_i] * ( person.E_ubo[hour_i] + ( person.E_sfo[hour_i] / ( 1- Beta_s)) + ( person.E_vfo[hour_i] / (1 - Beta_ev)))
        R_profit = R_profit + ParaOfRTP[DAY][hour_i] * ( person.E_uus[hour_i] + person.E_uuv[hour_i]) + ParaOfMarketP[DAY][hour_i] * (person.E_sso[hour_i] + person.E_vso[hour_i])
    R_capital = (( ParaOfMarketP.sum(axis = 1)[DAY]) + ( ParaOfRTP.sum(axis = 1)[DAY])) * (person.check_s_on_t[-1] + person.check_ev_on_t[-1]) / 48
    Fitness = C_grid + C_user - R_profit - R_capital
    
    return Fitness

def ChechLegalOrNot( person):
    if min(person.check_s_on_t) < 0:
        print( "Storage illegal")
    if min(person.check_ev_on_t) < 0:
        print( "EV illegal")


def sigmoid( getNumpy):
    return ( 1/( 1 + np.exp(np.negative(getNumpy))))




    



# Step1:產生起始解
population = list()
velocity = list()
HistoryBestInPast = list()

GlobalBest = list()

for i in range( NumOfPerson):
    population.append( person())


# 起始值
for i in range( NumOfPerson):
    
    # 產生人口前先初始化 
    tmp_E_uug, tmp_E_sfg, tmp_E_uus = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_E_sso, tmp_E_sfo, tmp_E_ubo = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_E_vfg, tmp_E_uuv, tmp_E_vfs = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_E_vso, tmp_E_vfo, tmp_theta_s, tmp_theta_ev = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
    tmp_s_on_t, tmp_ev_on_t = np.zeros((1, Hours + 1)), np.zeros((1, Hours + 1))
    
    if DAY == 0:
        tmp_s_on_t[0][0] = E_s_min
        tmp_ev_on_t[0][0] = E_ev_min
    else:
        file = load_workbook( PathOfFile + 'Remain0416.xlsx')
        sheet_ranges = file['HowManyStillHave_PSO']
    
        for k in range( len( DAY_List)):
            if DAY == list(enumerate(DAY_List))[k][0]:
                tmp_s_on_t[0][0] = (sheet_ranges.cell( row = k + 1, column = 2).value)
                tmp_ev_on_t[0][0] = (sheet_ranges.cell( row = k + 1, column = 3).value)
                break
    
    excel_s_on_t = tmp_s_on_t[0][0]
    excel_ev_on_t = tmp_ev_on_t[0][0]
    
    population[i].check_s_on_t.append( tmp_s_on_t[0][0])
    population[i].check_ev_on_t.append( tmp_ev_on_t[0][0])
    
    for hour_i in range( Hours):
        
        # 確認儲電設備
        if tmp_s_on_t[0][hour_i] <= E_s_min:                  # 儲電設備沒存量不能放電
            tmp_theta_s[0][hour_i] = 1
        elif tmp_s_on_t[0][hour_i] + ParaOfRenergy[DAY][hour_i] >= E_s_max:   # 儲電設備沒容量不能儲電
            tmp_theta_s[0][hour_i] = 0
        else:
            tmp_theta_s[0][hour_i] = random.randint(0, 1)
        
        # 確認電動
        if tmp_ev_on_t[0][hour_i] <= E_ev_min:
            tmp_theta_ev[0][hour_i] = 1
        elif tmp_ev_on_t[0][hour_i] >= E_ev_max or Theta_Work[DAY][hour_i] == 1:
            tmp_theta_ev[0][hour_i] = 0
        else:
            tmp_theta_ev[0][hour_i] = random.randint(0, 1)
        
        # 產生儲放電值
        
        # 當儲電設備儲電，電動車儲電
        if tmp_theta_s[0][hour_i] == 1 and tmp_theta_ev[0][hour_i] == 1:            
            
            # 儲電設備
            tmp_E_sfg[0][hour_i] = random.uniform(0, min( (E_s_max - tmp_s_on_t[0][hour_i] - ParaOfRenergy[DAY][hour_i]), Rate_s))
            tmp_E_sfo[0][hour_i] = random.uniform(0, min( max(E_s_max - tmp_s_on_t[0][hour_i] - tmp_E_sfg[0][hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s))        
            
            #電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_vfg[0][hour_i] = random.uniform(0, min(E_ev_max - tmp_ev_on_t[0][hour_i], Rate_ev))
                tmp_E_vfo[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfg[0][hour_i], 0), Rate_ev))
        
        # 當儲電設備放電，電動車儲電
        elif tmp_theta_s[0][hour_i] == 0 and tmp_theta_ev[0][hour_i] == 1:
            
            # 儲電設備
            tmp_E_uus[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min( tmp_s_on_t[0][hour_i], Rate_s)))
            tmp_E_sso[0][hour_i] = random.uniform(0, min(max( tmp_s_on_t[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s))
            
            #電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_vfs[0][hour_i] = random.uniform(0, min(E_ev_max - tmp_ev_on_t[0][hour_i], min(max( tmp_s_on_t[0][hour_i] - tmp_E_sso[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s)))
                tmp_E_vfg[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfs[0][hour_i],0), Rate_ev))
                tmp_E_vfo[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfg[0][hour_i] - tmp_E_vfs[0][hour_i], 0), Rate_ev))
        
        # 當儲電設備儲電，電動車放電
        elif tmp_theta_s[0][hour_i] == 1 and tmp_theta_ev[0][hour_i] == 0:
        
            # 儲電設備
            tmp_E_sfg[0][hour_i] = random.uniform(0, min( (E_s_max - tmp_s_on_t[0][hour_i] - ParaOfRenergy[DAY][hour_i]), Rate_s))
            tmp_E_sfo[0][hour_i] = random.uniform(0, min( max(E_s_max - tmp_s_on_t[0][hour_i] - tmp_E_sfg[0][hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s))
            
            #儲電設備
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_uuv[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min(tmp_ev_on_t[0][hour_i] - E_ev_min, Rate_ev)))
                tmp_E_vso[0][hour_i] = random.uniform(0, min(max(tmp_ev_on_t[0][hour_i] - tmp_E_uuv[0][hour_i] - E_ev_min, 0), Rate_ev))
        
        # 當儲電設備放電，電動車放電
        elif tmp_theta_s[0][hour_i] == 0 and tmp_theta_ev[0][hour_i] == 0:
        
            # 儲電設備
            tmp_E_uus[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min( tmp_s_on_t[0][hour_i], Rate_s)))
            tmp_E_sso[0][hour_i] = random.uniform(0, min(max( tmp_s_on_t[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s))
            # 電動車
            if Theta_Work[DAY][hour_i] == 0:
                tmp_E_uuv[0][hour_i] = random.uniform(0, min(max(0, D_user[0][hour_i] - tmp_E_uus[0][hour_i]), min(tmp_ev_on_t[0][hour_i] - E_ev_min, Rate_ev)))
                tmp_E_vso[0][hour_i] = random.uniform(0, min(max(tmp_ev_on_t[0][hour_i] - tmp_E_uuv[0][hour_i] - E_ev_min, 0), Rate_ev))
        else:
            print('Error in initial!!')
        
        # 隨機產生E_uug
        tmp_E_uug[0][hour_i] = random.uniform(0, max(D_user[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_uus[0][hour_i], 0))
        
        # E_ubo必須用來滿足D_user
        tmp_E_ubo[0][hour_i] = max(D_user[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_uus[0][hour_i] - tmp_E_uug[0][hour_i], 0)
        tmp_s_on_t[0][hour_i+1] = tmp_s_on_t[0][hour_i] + tmp_E_sfg[0][hour_i] + tmp_E_sfo[0][hour_i] - tmp_E_uus[0][hour_i] - tmp_E_sso[0][hour_i] - tmp_E_vfs[0][hour_i] + ParaOfRenergy[DAY][hour_i]
        tmp_ev_on_t[0][hour_i+1] = tmp_ev_on_t[0][hour_i] + tmp_E_vfg[0][hour_i] + tmp_E_vfo[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_vso[0][hour_i]
        
        # 將產生候選解加到人口中
        population[i].E_uug.append(tmp_E_uug[0][hour_i])
        population[i].E_sfg.append(tmp_E_sfg[0][hour_i])
        population[i].E_uus.append(tmp_E_uus[0][hour_i])
        population[i].E_sso.append(tmp_E_sso[0][hour_i])
        population[i].E_sfo.append(tmp_E_sfo[0][hour_i])
        population[i].E_ubo.append(tmp_E_ubo[0][hour_i])
        population[i].E_vfg.append(tmp_E_vfg[0][hour_i])
        population[i].E_uuv.append(tmp_E_uuv[0][hour_i])
        population[i].E_vfs.append(tmp_E_vfs[0][hour_i])
        population[i].E_vso.append(tmp_E_vso[0][hour_i])
        population[i].E_vfo.append(tmp_E_vfo[0][hour_i])
        population[i].theta_s.append(tmp_theta_s[0][hour_i])
        population[i].theta_ev.append(tmp_theta_ev[0][hour_i])
        population[i].check_s_on_t.append( tmp_s_on_t[0][hour_i +1])
        population[i].check_ev_on_t.append( tmp_ev_on_t[0][hour_i +1])
    
        # 將List轉成矩陣格式以利計算
        population[i].matrix[0][hour_i] = tmp_E_uug[0][hour_i]
        population[i].matrix[1][hour_i] = tmp_E_sfg[0][hour_i]
        population[i].matrix[2][hour_i] = tmp_E_uus[0][hour_i]
        population[i].matrix[3][hour_i] = tmp_E_sso[0][hour_i]
        population[i].matrix[4][hour_i] = tmp_E_sfo[0][hour_i]
        population[i].matrix[5][hour_i] = tmp_E_ubo[0][hour_i]
        population[i].matrix[6][hour_i] = tmp_E_vfg[0][hour_i]
        population[i].matrix[7][hour_i] = tmp_E_uuv[0][hour_i]
        population[i].matrix[8][hour_i] = tmp_E_vfs[0][hour_i]
        population[i].matrix[9][hour_i] = tmp_E_vso[0][hour_i]
        population[i].matrix[10][hour_i] = tmp_E_vfo[0][hour_i]
        population[i].matrix[11][hour_i] = tmp_theta_s[0][hour_i]
        population[i].matrix[12][hour_i] = tmp_theta_ev[0][hour_i]
    
    # 各自過去最佳的解存起來
    population[i].previousBestMatrix = population[i].matrix

# Step1.5:檢查是否有不可行解    
for i in range( NumOfPerson):
    ChechLegalOrNot( population[i])
    
# Step2:計算適應度值    
for i in range( NumOfPerson):
    population[i].fitness = calTheFit( population[i])
    population[i].previousFitness = population[i].fitness
    
# Step2.5:存取t_0世代下最好的解
FitnessArray = list()                                               # 用來存當世代所有候選解的fitness
for i in range( NumOfPerson):
    FitnessArray.append(population[i].fitness)                      # temp矩陣存當世代所有人的fitness
    
IndexOfBestCS = FitnessArray.index( min(FitnessArray))              # 找到第0世代最好的解 並返回索引值
BestCSmatrix = population[IndexOfBestCS].matrix

Hist_BestFit.append( population[IndexOfBestCS].fitness)
Hist_PreBest.append( population[IndexOfBestCS].fitness)
Hist_AvgFit.append( sum( FitnessArray)/len( FitnessArray))

tmp_check_s_on_t = population[IndexOfBestCS].check_s_on_t[-1]
tmp_check_ev_on_t = population[IndexOfBestCS].check_ev_on_t[-1]
            


# 初始化速度
for i in range( NumOfPerson):  
    population[i].velocity = np.random.uniform( -Par_SetRange, Par_SetRange, size = ( 13, Hours))

time_iter_start = time.time()
# Step3:主迴圈開始    
for iteration in range( MaxOfIteration):
    
    
    
    OffSpring = list()                                  # 宣告子代
    for i in range(NumOfPerson):                        
        OffSpring.append(person())                      
        
        # 初始化儲電設備與電動車每小時的電量 
        tmp_E_uug, tmp_E_sfg, tmp_E_uus = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_E_sso, tmp_E_sfo, tmp_E_ubo = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_E_vfg, tmp_E_uuv, tmp_E_vfs = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_E_vso, tmp_E_vfo, tmp_theta_s, tmp_theta_ev = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_s_on_t, tmp_ev_on_t = np.zeros((1, Hours + 1)), np.zeros((1, Hours + 1))
    
    
        if DAY == 0:
            tmp_s_on_t[0][0] = E_s_min
            tmp_ev_on_t[0][0] = E_ev_min
            
        else:
#            file = load_workbook( PathOfFile + 'Remain0416.xlsx')
#            sheet_ranges = file['HowManyStillHave_PSO']
#            
#            for k in range( len( DAY_List)):
#                if DAY == list(enumerate(DAY_List))[k][0]:
#                    tmp_s_on_t[0][0] = sheet_ranges.cell( row = k + 1, column = 2).value
#                    tmp_ev_on_t[0][0] = sheet_ranges.cell( row = k + 1, column = 3).value
#                    break
            
            tmp_s_on_t[0][0] = excel_s_on_t
            tmp_ev_on_t[0][0] = excel_ev_on_t
            
            
        
        
        OffSpring[i].check_s_on_t.append( tmp_s_on_t[0][0])
        OffSpring[i].check_ev_on_t.append( tmp_ev_on_t[0][0])

        
        
        
        # ========================== ============================
        # 更新速度
        
        rand_1 = np.random.random()
        rand_2 = np.random.random()
        
        tmpNewV = Par_w * population[i].velocity + Par_c1 * rand_1 * ( population[i].previousBestMatrix - population[i].matrix) + Par_c2 * rand_2 * ( BestCSmatrix - population[i].matrix)
        
        # 檢查是否超過速度上限值
        for check_r in range( tmpNewV.shape[0]):
            for check_c in range( tmpNewV.shape[1]):
                if tmpNewV[check_r][check_c] > Par_Vmax:
                    
                    tmpNewV[check_r][check_c] = Par_Vmax
        
        OffSpring[i].velocity = tmpNewV
            
        # 二元變數使用BPSO找鄰近解
        tmpRand = np.random.uniform( 0, 1, size = ( 2, Hours))
           
        tmpToCompare = sigmoid( tmpNewV[11:13][:])
                
        for check_r in range( tmpRand.shape[0]):
            for check_c in range( tmpRand.shape[1]):
                
                if DAY != 5 or DAY != 6:
                    if tmpRand[check_r][check_c] < tmpToCompare[check_r][check_c]:
                        if check_r == 1 and check_c > 6 and check_c < 17:
                                tmpToCompare[check_r][check_c] = 0
                        else:
                            tmpToCompare[check_r][check_c] = 1
                    else:
                        tmpToCompare[check_r][check_c] = 0
                else:
                    if tmpRand[check_r][check_c] < tmpToCompare[check_r][check_c]:
                        tmpToCompare[check_r][check_c] = 1
                    else:
                        tmpToCompare[check_r][check_c] = 0
        
        
        tmpNewLocation = population[i].matrix + tmpNewV
        
        for check_r in range( tmpRand.shape[0]):
            for check_c in range( tmpRand.shape[1]):
                tmpNewLocation[check_r + 11][check_c] = tmpToCompare[check_r][check_c]
        
        # 若有負號 則修正至0
        for check_r in range( tmpNewLocation.shape[0]):
            for check_c in range( tmpNewLocation.shape[1]):
                tmpNewLocation[check_r][check_c] = max( 0, tmpNewLocation[check_r][check_c])
        
        
        # 鄰近解變更完成
        for hour_i in range( Hours):
            OffSpring[i].E_uug.append(tmpNewLocation[0][hour_i])
            OffSpring[i].E_sfg.append(tmpNewLocation[1][hour_i])
            OffSpring[i].E_uus.append(tmpNewLocation[2][hour_i])
            OffSpring[i].E_sso.append(tmpNewLocation[3][hour_i])
            OffSpring[i].E_sfo.append(tmpNewLocation[4][hour_i])
            OffSpring[i].E_ubo.append(tmpNewLocation[5][hour_i])
            OffSpring[i].E_vfg.append(tmpNewLocation[6][hour_i])
            OffSpring[i].E_uuv.append(tmpNewLocation[7][hour_i])
            OffSpring[i].E_vfs.append(tmpNewLocation[8][hour_i])
            OffSpring[i].E_vso.append(tmpNewLocation[9][hour_i])
            OffSpring[i].E_vfo.append(tmpNewLocation[10][hour_i])
            OffSpring[i].theta_s.append(tmpNewLocation[11][hour_i])
            OffSpring[i].theta_ev.append(tmpNewLocation[12][hour_i])
        
        
    # Step4:修正不可行解
    for i in range( NumOfPerson):
        
        count_fix = 0
        
        OffSpring[i].check_s_on_t = list()
        OffSpring[i].check_ev_on_t = list()
        
        if DAY == 0:
            OffSpring[i].check_s_on_t.append(E_s_min)
            OffSpring[i].check_ev_on_t.append(E_ev_min) 
        else:
#            file = load_workbook( PathOfFile + 'Remain0416.xlsx')
#            sheet_ranges = file['HowManyStillHave_PSO']
#            
#            for k in range( len( DAY_List)):
#                if DAY == list(enumerate(DAY_List))[k][0]:
#                    OffSpring[i].check_s_on_t.append( sheet_ranges.cell( row = k + 1, column = 2).value)
#                    OffSpring[i].check_ev_on_t.append( sheet_ranges.cell( row = k + 1, column = 3).value) 
#                    break

            OffSpring[i].check_s_on_t.append( excel_s_on_t)
            OffSpring[i].check_ev_on_t.append( excel_ev_on_t) 

            
        for hour_i in range( Hours):
            OffSpring[i].check_s_on_t.append( 0)
            OffSpring[i].check_ev_on_t.append( 0)
        
        for hour_i in range( Hours):
            
            # 1. 檢查儲放電設備是否有放電錯誤
            if OffSpring[i].theta_s[hour_i] == 0:
                if OffSpring[i].E_sfg[hour_i] > 0:
                    OffSpring[i].E_sfg[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_sfo[hour_i] > 0:
                    OffSpring[i].E_sfo[hour_i] = 0
                    count_fix = count_fix + 1
            else:
                if OffSpring[i].E_uus[hour_i] > 0:
                    OffSpring[i].E_uug[hour_i] = OffSpring[i].E_uug[hour_i] + OffSpring[i].E_uus[hour_i]
                    OffSpring[i].E_uus[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_sso[hour_i] > 0:
                    OffSpring[i].E_sso[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_vfs[hour_i] > 0:
                    OffSpring[i].E_vfg[hour_i] = OffSpring[i].E_vfg[hour_i] + OffSpring[i].E_vfs[hour_i]
                    OffSpring[i].E_vfs[hour_i] = 0
                    count_fix = count_fix + 1
            # 2. 檢查電動車是否由儲放電錯誤
            # 如果在上班時段
            if Theta_Work[DAY][hour_i] == 1:
                if OffSpring[i].E_vfg[hour_i] > 0:
                    OffSpring[i].E_vfg[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_vfs[hour_i] > 0:
                    OffSpring[i].E_sso[hour_i] = OffSpring[i].E_sso[hour_i] + OffSpring[i].E_vfs[hour_i]
                    OffSpring[i].E_vfs[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_vfo[hour_i] > 0:
                    OffSpring[i].E_vfo[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_uuv[hour_i] > 0:
                    OffSpring[i].E_uug[hour_i] = OffSpring[i].E_uuv[hour_i] + OffSpring[i].E_uuv[hour_i]
                    OffSpring[i].E_uuv[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_vso[hour_i] > 0:
                    OffSpring[i].E_vso[hour_i] = 0
                    count_fix = count_fix + 1
    
            if OffSpring[i].theta_ev[hour_i] == 0:
                if OffSpring[i].E_vfg[hour_i] > 0:
                    OffSpring[i].E_vfg[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_vfs[hour_i] > 0:
                    OffSpring[i].E_sso[hour_i] = OffSpring[i].E_sso[hour_i] + OffSpring[i].E_vfs[hour_i]
                    OffSpring[i].E_vfs[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_vfo[hour_i] > 0:
                    OffSpring[i].E_vfo[hour_i] = 0
                    count_fix = count_fix + 1
            else:
                if OffSpring[i].E_uuv[hour_i] > 0:
                    OffSpring[i].E_uug[hour_i] = OffSpring[i].E_uug[hour_i] + OffSpring[i].E_uuv[hour_i]
                    OffSpring[i].E_uuv[hour_i] = 0
                    count_fix = count_fix + 1
                if OffSpring[i].E_vso[hour_i] > 0:
                    OffSpring[i].E_vso[hour_i] = 0
                    count_fix = count_fix + 1
            # 3. 檢查儲電設備與電動車容量限制
            tmp_s = OffSpring[i].check_s_on_t[hour_i] + OffSpring[i].E_sfg[hour_i] + OffSpring[i].E_sfo[hour_i] - OffSpring[i].E_uus[hour_i] - OffSpring[i].E_sso[hour_i] - OffSpring[i].E_vfs[hour_i] + ParaOfRenergy[DAY][hour_i]
            tmp_ev = OffSpring[i].check_ev_on_t[hour_i] + OffSpring[i].E_vfg[hour_i] + OffSpring[i].E_vfs[hour_i] + OffSpring[i].E_vfo[hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_vso[hour_i]
            
            
            
            # 當儲電設備發生沒電還放電或是滿載還儲電的話, 就要修正
            if ((OffSpring[i].check_s_on_t[ hour_i + 1] < E_s_min) or (OffSpring[i].check_s_on_t[ hour_i + 1] > E_s_max)) or ((OffSpring[i].check_ev_on_t[ hour_i + 1] < E_ev_min) or ( OffSpring[i].check_ev_on_t[ hour_i + 1] > E_ev_max)):
                
                count_fix = count_fix + 1
                OffSpring[i].E_uug[hour_i] = 0
                OffSpring[i].E_sfg[hour_i] = 0
                OffSpring[i].E_uus[hour_i] = 0
                OffSpring[i].E_sso[hour_i] = 0
                OffSpring[i].E_sfo[hour_i] = 0
                OffSpring[i].E_ubo[hour_i] = 0
                OffSpring[i].E_vfg[hour_i] = 0
                OffSpring[i].E_uuv[hour_i] = 0
                OffSpring[i].E_vfs[hour_i] = 0
                OffSpring[i].E_vso[hour_i] = 0
                OffSpring[i].E_vfo[hour_i] = 0
                
                # 儲電設備沒存量不能放電
                if OffSpring[i].check_s_on_t[ hour_i] <= E_s_min:
                    OffSpring[i].theta_s[hour_i] = 1
                    count_fix = count_fix + 1
                # 儲電設備沒容量不能儲電
                elif OffSpring[i].check_s_on_t[ hour_i] + ParaOfRenergy[DAY][hour_i] >= E_s_max:
                    OffSpring[i].theta_s[hour_i] = 0
                    count_fix = count_fix + 1
                else:
                    OffSpring[i].theta_s[hour_i] = np.random.randint(0, 2)
                    count_fix = count_fix + 1
                # 電動車沒存量不能放電
                if OffSpring[i].check_ev_on_t[ hour_i] <= E_ev_min:
                    OffSpring[i].theta_ev[hour_i] = 1
                    count_fix = count_fix + 1
                # 電動車沒容量不能儲電
                elif OffSpring[i].check_ev_on_t[ hour_i] >= E_ev_max:
                    OffSpring[i].theta_ev[hour_i] = 0
                    count_fix = count_fix + 1
                else:
                    OffSpring[i].theta_ev[hour_i] = np.random.randint(0, 2)
                    count_fix = count_fix + 1
                # 儲電設備及電動車同時儲電情況
                if OffSpring[i].theta_s[hour_i] == 1 and OffSpring[i].theta_ev[hour_i] == 1:
                    count_fix = count_fix + 1
                    # 儲電設備
                    OffSpring[i].E_sfg[hour_i] = random.uniform( 0, min( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - ParaOfRenergy[DAY][hour_i], Rate_s) )
                    OffSpring[i].E_sfo[hour_i] = random.uniform( 0, min( max( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_sfg[hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s ))
                    # 電動車
                    if Theta_Work[DAY][hour_i] == 0:
                        OffSpring[i].E_vfg[hour_i] = random.uniform( 0, min( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i], Rate_ev))
                        OffSpring[i].E_vfo[hour_i] = random.uniform( 0, min( max( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_vfg[hour_i], 0), Rate_ev))
                
                # 產生儲放電值，當儲電設備放電，電動車儲電
                elif OffSpring[i].theta_s[hour_i] == 0 and OffSpring[i].theta_ev[hour_i] == 1:
                    count_fix = count_fix + 1
                    # 儲電設備
                    OffSpring[i].E_uus[hour_i] = random.uniform( 0, min( D_user[0][hour_i], min( OffSpring[i].check_s_on_t[ hour_i], Rate_s)))
                    OffSpring[i].E_sso[hour_i] = random.uniform( 0, min( max( OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_uus[hour_i], 0), Rate_s))
                    # 電動車
                    if Theta_Work[DAY][hour_i] == 0:
                        OffSpring[i].E_vfs[hour_i] = random.uniform( 0, min( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i], min( max(  OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_sso[hour_i] - OffSpring[i].E_uus[hour_i],0), Rate_s)))
                        OffSpring[i].E_vfg[hour_i] = random.uniform( 0, min( max( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_vfs[hour_i], 0), Rate_ev))
                        OffSpring[i].E_vfo[hour_i] = random.uniform( 0, min( max( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_vfg[hour_i] - OffSpring[i].E_vfs[hour_i], 0), Rate_ev))
                
                # 產生儲放電值，當儲電設備儲電，電動車放電
                elif OffSpring[i].theta_s[hour_i] == 1 and OffSpring[i].theta_ev[hour_i] == 0:
                    count_fix = count_fix + 1
                    # 儲電設備
                    OffSpring[i].E_sfg[hour_i] = random.uniform(0, min( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - ParaOfRenergy[DAY][hour_i], Rate_s))
                    OffSpring[i].E_sfo[hour_i] = random.uniform( 0, min( max( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_sfg[hour_i] - ParaOfRenergy[DAY][hour_i], 0), Rate_s ))
                    # 電動車
                    if Theta_Work[DAY][hour_i] == 0:
                        OffSpring[i].E_uuv[hour_i] = random.uniform(0, min( D_user[0][hour_i], min( OffSpring[i].check_ev_on_t[ hour_i] - E_ev_min, Rate_ev)))
                        OffSpring[i].E_vso[hour_i] = random.uniform(0, min(max( OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_uuv[hour_i] - E_ev_min, 0), Rate_ev))
                
                # 產生儲放電值，當儲電設備放電，電動車放電
                else:
                    count_fix = count_fix + 1
                    # 儲電設備
                    OffSpring[i].E_uus[hour_i] = random.uniform( 0, min( D_user[0][hour_i], min( OffSpring[i].check_s_on_t[ hour_i], Rate_s)))
                    OffSpring[i].E_sso[hour_i] = random.uniform( 0, min( max( OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_uus[hour_i], 0), Rate_s))
                    # 電動車
                    if Theta_Work[DAY][hour_i] == 0:
                        OffSpring[i].E_uuv[hour_i] = random.uniform( 0, min( max( 0, D_user[0][hour_i] - OffSpring[i].E_uus[hour_i]), min( OffSpring[i].check_ev_on_t[ hour_i] - E_ev_min, Rate_ev)))
                        OffSpring[i].E_vso[hour_i] = random.uniform( 0, min( max( OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_uuv[hour_i] - E_ev_min, 0), Rate_ev))
                
                # 隨機產生E_uug
                OffSpring[i].E_uug[hour_i] = random.uniform( 0, max( D_user[0][hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_uus[hour_i], 0))
                # E_ubo必須用來滿足D_user
                OffSpring[i].E_ubo[hour_i] = max( D_user[0][hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_uus[hour_i] - OffSpring[i].E_uug[hour_i], 0)
                    
            # 更新結果
            tmp_s = OffSpring[i].check_s_on_t[hour_i] + OffSpring[i].E_sfg[hour_i] + OffSpring[i].E_sfo[hour_i] - OffSpring[i].E_uus[hour_i] - OffSpring[i].E_sso[hour_i] - OffSpring[i].E_vfs[hour_i] + ParaOfRenergy[DAY][hour_i]
            tmp_ev = OffSpring[i].check_ev_on_t[hour_i] + OffSpring[i].E_vfg[hour_i] + OffSpring[i].E_vfs[hour_i] + OffSpring[i].E_vfo[hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_vso[hour_i]
                        
            OffSpring[i].check_s_on_t[hour_i +1] = tmp_s
            OffSpring[i].check_ev_on_t[hour_i +1] = tmp_ev
    
#        print( '第', iteration+1, '迭代下的', i, '條鯨魚修正了', count_fix)
    
    # Step5:檢查修正後是否有有不可行解        
    for i in range( NumOfPerson):
        ChechLegalOrNot( OffSpring[i])
     
    # Step6:計算適應度值    
    for i in range( NumOfPerson):
        OffSpring[i].fitness = calTheFit( OffSpring[i])            
    
        # 與自己比較, 若匹過去最佳來的好, 就替換過去最佳解, 若無, 則不更動
        
        for hour_i in range( Hours):
            OffSpring[i].matrix[0][hour_i] = OffSpring[i].E_uug[hour_i]
            OffSpring[i].matrix[1][hour_i] = OffSpring[i].E_sfg[hour_i]
            OffSpring[i].matrix[2][hour_i] = OffSpring[i].E_uus[hour_i]
            OffSpring[i].matrix[3][hour_i] = OffSpring[i].E_sso[hour_i]
            OffSpring[i].matrix[4][hour_i] = OffSpring[i].E_sfo[hour_i]
            OffSpring[i].matrix[5][hour_i] = OffSpring[i].E_ubo[hour_i]
            OffSpring[i].matrix[6][hour_i] = OffSpring[i].E_vfg[hour_i]
            OffSpring[i].matrix[7][hour_i] = OffSpring[i].E_uuv[hour_i]
            OffSpring[i].matrix[8][hour_i] = OffSpring[i].E_vfs[hour_i]
            OffSpring[i].matrix[9][hour_i] = OffSpring[i].E_vso[hour_i]
            OffSpring[i].matrix[10][hour_i] = OffSpring[i].E_vfo[hour_i]
            OffSpring[i].matrix[11][hour_i] = OffSpring[i].theta_s[hour_i]
            OffSpring[i].matrix[12][hour_i] = OffSpring[i].theta_ev[hour_i]
        
#        print(OffSpring[i].fitness)
        # 若比過去最佳好
        if OffSpring[i].fitness <= population[i].previousFitness:
            OffSpring[i].previousFitness = OffSpring[i].fitness
            OffSpring[i].previousBestMatrix = OffSpring[i].matrix
            
            
        # 若沒有比較好, 則繼續保留之前的候選解與其適應度值
        else:
            OffSpring[i].previousFitness = population[i].previousFitness
            OffSpring[i].previousBestMatrix = population[i].previousBestMatrix
    
    
    # 開始移動
    
    population = list()
    for i in range( len(OffSpring)):
        population.append(OffSpring[i])
    
    # 將最佳解存起來
    FitnessArray = list()                                                   # 用來存當世代所有候選解的fitness
    FitnessInPrevious = list()                                              # 用來存取歷史嘴加的fitness
    for i in range( NumOfPerson):
        FitnessArray.append( population[i].fitness)                        
        FitnessInPrevious.append( population[i].previousFitness)
    
    IndexOfBestCS = FitnessInPrevious.index( min(FitnessInPrevious))        # 找到過去最好的適應度值 並返回索引值
    BestCSmatrix = population[IndexOfBestCS].previousBestMatrix
    
    
    Hist_BestFit.append( population[FitnessArray.index( min(FitnessArray))].fitness)
    Hist_PreBest.append( population[IndexOfBestCS].previousFitness)
    
    if Hist_PreBest[-1] < Hist_PreBest[-2]:
        tmp_check_s_on_t =population[IndexOfBestCS].check_s_on_t[-1]
        tmp_check_ev_on_t = population[IndexOfBestCS].check_ev_on_t[-1]
            
    
    
    Hist_AvgFit.append( sum( FitnessArray)/len( FitnessArray))
    
    
    time_iter_end = time.time()
    
#        time_FLAG = time.time()
    
    
    if iteration % 100 == 0:
        print( 'now in', iteration, 'iteration')
        print( 'one iteration spent:', time_iter_end - time_iter_start )
        time_iter_start = time.time()
#            print( "now are spent :", round(time_FLAG - time_start, 4))
    
#    if (( time_FLAG - time_start) >= 825):
#        break
    

# ==== 演算法結束 ====
# Step8:看最後的解
BestCSSolution = BestCSmatrix

# 將儲電設備與電動車最後的電量儲存起來
ToTxt_s = tmp_check_s_on_t
ToTxt_ev = tmp_check_s_on_t

file = load_workbook( PathOfFile + 'Remain0416.xlsx')
sheet_ranges = file['HowManyStillHave_PSO']

for k in range( len( DAY_List)):
    if DAY == list(enumerate(DAY_List))[k][0]:
        sheet_ranges.cell(row = k + 2, column = 2).value = ToTxt_s
        sheet_ranges.cell(row = k + 2, column = 3).value = ToTxt_ev
        break

file.save( 'Remain0416.xlsx')
file.close()



   
    
    
