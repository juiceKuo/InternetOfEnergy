"""
Created on MON 0204 22:04:46 2019
@author: ChihChi
"""
import pandas as pd
import numpy as np
import random
import time
from openpyxl import Workbook
from openpyxl import load_workbook




List_bestValue = list()
List_time = list()


for iRun in range( 5):

    #參數設定
    NumOfPerson = 20
    Hours = 24
    E_s_min = 0
    E_ev_min = 12.5
    E_s_max = 13.5
    E_ev_max = 50
    Rate_s = 5
    Rate_ev = 10 
    Beta_s = 0.1
    Beta_ev = 0.1
    
    # Par微調參數
    Interval = 0.5
    
    # 目前週幾
    # DAY = 0 -> Monday, DAY = 1 -> Tuesday, DAY = 2 -> Wed
    DAY = 0
    DAY_List = [ 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT', 'SUN']
    
    Hist_BestFit = list()
    Hist_AvgFit = list()
    #Hist_forCompare = list()
    
    
    # 演算法參數
    MaxOfIteration = 25000
    Par_HMCR = 0.9
    Par_PAR = 0.5 
    Rate_s = 5
    Rate_ev = 8
    
    
    a = random.randint( 1, 150000)
#    a = 106076
    
    # 設定亂數種子
    random.seed( a)
    np.random.seed( a)
    
    
    # 建立物件
    class person:
        def __init__(self):
            self.E_uug = list()
            self.E_sfg = list()
            self.E_uus = list()
            self.E_sso = list()
            self.E_sfo = list()
            self.E_ubo = list()
            self.E_vfg = list()
            self.E_uuv = list()
            self.E_vfs = list()
            self.E_vso = list()
            self.E_vfo = list()
            self.theta_s = list()
            self.theta_ev = list()
            self.check_s_on_t = list()
            self.check_ev_on_t = list()
            self.fitness = 0
            
    # =============================================================================        
    PathOfFile = 'C:/Users/juice/Desktop/MyProposal/IoE_Program/'
    
    # == 可再生能源 ==
    target = pd.read_excel( PathOfFile + 'RenewableEnergy.xlsx', dtype = 'float64')
    ParaOfRenergy = target.values
    
    # == 實時電價 ==
    target = pd.read_excel( PathOfFile + 'RTP.xlsx', dtype = 'float64')
    ParaOfRTP = target.values
    
    # == 市場價格 ==
    target = pd.read_excel( PathOfFile + 'MarketingPrice.xlsx', dtype = 'float64')
    ParaOfMarketP = target.values
    
    # == D_user ==
    target = pd.read_excel( PathOfFile + 'Duser.xlsx', dtype = 'float64')
    D_user = target.values
    
    # == Theta_Work ==
    target = pd.read_excel( PathOfFile + 'Theta_work.xlsx', dtype = 'float64')
    Theta_Work = target.values
    
    # =============================================================================
    
    # function
    def calTheFit( person):
        
        C_grid, C_user, R_profit, R_capital = 0, 0, 0, 0
        for hour_i in range( Hours):
            C_grid = C_grid + ParaOfRTP[DAY][hour_i] * ( person.E_uug[hour_i] + ( person.E_sfg[hour_i] / ( 1 - Beta_s)) + ( person.E_vfg[hour_i] / (1 - Beta_ev)))
            C_user = C_user + ParaOfMarketP[DAY][hour_i] * ( person.E_ubo[hour_i] + ( person.E_sfo[hour_i] / ( 1- Beta_s)) + ( person.E_vfo[hour_i] / (1 - Beta_ev)))
            R_profit = R_profit + ParaOfRTP[DAY][hour_i] * ( person.E_uus[hour_i] + person.E_uuv[hour_i]) + ParaOfMarketP[DAY][hour_i] * (person.E_sso[hour_i] + person.E_vso[hour_i])
        R_capital = (( ParaOfMarketP.sum(axis = 1)[DAY]) + ( ParaOfRTP.sum(axis = 1)[DAY])) * (person.check_s_on_t[-1] + person.check_ev_on_t[-1]) / 48
        Fitness = C_grid + C_user - R_profit - R_capital
        
        return Fitness
    
    def ChechLegalOrNot( person):
        if min(person.check_s_on_t) < 0:
            print( "Storage illegal")
        if min(person.check_ev_on_t) < 0:
            print( "EV illegal")
    
    
    
    
    time_start = time.time()
    
    
    # Step1:產生起始解    
    population = list()
    for i in range( NumOfPerson):
        population.append( person())
        
    # 起始值
    for i in range( NumOfPerson):
        
        # 產生人口前先初始化 
        tmp_E_uug, tmp_E_sfg, tmp_E_uus = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_E_sso, tmp_E_sfo, tmp_E_ubo = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_E_vfg, tmp_E_uuv, tmp_E_vfs = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_E_vso, tmp_E_vfo, tmp_theta_s, tmp_theta_ev = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
        tmp_s_on_t, tmp_ev_on_t = np.zeros((1, Hours + 1)), np.zeros((1, Hours + 1))
        
        if DAY == 0:
            tmp_s_on_t[0][0] = E_s_min
            tmp_ev_on_t[0][0] = E_ev_min
        else:
            file = load_workbook( PathOfFile + 'Remain0416.xlsx')
            sheet_ranges = file['HowManyStillHave_HSA']
        
            for k in range( len( DAY_List)):
                if DAY == list(enumerate(DAY_List))[k][0]:
                    tmp_s_on_t[0][0] = (sheet_ranges.cell( row = k + 1, column = 2).value)
                    tmp_ev_on_t[0][0] = (sheet_ranges.cell( row = k + 1, column = 3).value)
                    break
        
        population[i].check_s_on_t.append( tmp_s_on_t[0][0])
        population[i].check_ev_on_t.append( tmp_ev_on_t[0][0])
        
        for hour_i in range( Hours):
            
            # 確認儲電設備
            if tmp_s_on_t[0][hour_i] <= E_s_min:                  # 儲電設備沒存量不能放電
                tmp_theta_s[0][hour_i] = 1
            elif tmp_s_on_t[0][hour_i] + ParaOfRenergy[0][hour_i] >= E_s_max:   # 儲電設備沒容量不能儲電
                tmp_theta_s[0][hour_i] = 0
            else:
                tmp_theta_s[0][hour_i] = random.randint(0, 1)
            
            # 確認電動
            if tmp_ev_on_t[0][hour_i] <= E_ev_min:
                tmp_theta_ev[0][hour_i] = 1
            elif tmp_ev_on_t[0][hour_i] >= E_ev_max or Theta_Work[DAY][hour_i] == 1:
                tmp_theta_ev[0][hour_i] = 0
            else:
                tmp_theta_ev[0][hour_i] = random.randint(0, 1)
            
            # 產生儲放電值
            
            # 當儲電設備儲電，電動車儲電
            if tmp_theta_s[0][hour_i] == 1 and tmp_theta_ev[0][hour_i] == 1:            
                
                # 儲電設備
                tmp_E_sfg[0][hour_i] = random.uniform(0, min( (E_s_max - tmp_s_on_t[0][hour_i] - ParaOfRenergy[0][hour_i]), Rate_s))
                tmp_E_sfo[0][hour_i] = random.uniform(0, min( max(E_s_max - tmp_s_on_t[0][hour_i] - tmp_E_sfg[0][hour_i] - ParaOfRenergy[0][hour_i], 0), Rate_s))        
                
                #電動車
                if Theta_Work[DAY][hour_i] == 0:
                    tmp_E_vfg[0][hour_i] = random.uniform(0, min(E_ev_max - tmp_ev_on_t[0][hour_i], Rate_ev))
                    tmp_E_vfo[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfg[0][hour_i], 0), Rate_ev))
            
            # 當儲電設備放電，電動車儲電
            elif tmp_theta_s[0][hour_i] == 0 and tmp_theta_ev[0][hour_i] == 1:
                
                # 儲電設備
                tmp_E_uus[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min( tmp_s_on_t[0][hour_i], Rate_s)))
                tmp_E_sso[0][hour_i] = random.uniform(0, min(max( tmp_s_on_t[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s))
                
                #電動車
                if Theta_Work[DAY][hour_i] == 0:
                    tmp_E_vfs[0][hour_i] = random.uniform(0, min(E_ev_max - tmp_ev_on_t[0][hour_i], min(max( tmp_s_on_t[0][hour_i] - tmp_E_sso[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s)))
                    tmp_E_vfg[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfs[0][hour_i],0), Rate_ev))
                    tmp_E_vfo[0][hour_i] = random.uniform(0, min(max(E_ev_max - tmp_ev_on_t[0][hour_i] - tmp_E_vfg[0][hour_i] - tmp_E_vfs[0][hour_i], 0), Rate_ev))
            
            # 當儲電設備儲電，電動車放電
            elif tmp_theta_s[0][hour_i] == 1 and tmp_theta_ev[0][hour_i] == 0:
            
                # 儲電設備
                tmp_E_sfg[0][hour_i] = random.uniform(0, min( (E_s_max - tmp_s_on_t[0][hour_i] - ParaOfRenergy[0][hour_i]), Rate_s))
                tmp_E_sfo[0][hour_i] = random.uniform(0, min( max(E_s_max - tmp_s_on_t[0][hour_i] - tmp_E_sfg[0][hour_i] - ParaOfRenergy[0][hour_i], 0), Rate_s))
                
                #儲電設備
                if Theta_Work[DAY][hour_i] == 0:
                    tmp_E_uuv[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min(tmp_ev_on_t[0][hour_i] - E_ev_min, Rate_ev)))
                    tmp_E_vso[0][hour_i] = random.uniform(0, min(max(tmp_ev_on_t[0][hour_i] - tmp_E_uuv[0][hour_i] - E_ev_min, 0), Rate_ev))
            
            # 當儲電設備放電，電動車放電
            elif tmp_theta_s[0][hour_i] == 0 and tmp_theta_ev[0][hour_i] == 0:
            
                # 儲電設備
                tmp_E_uus[0][hour_i] = random.uniform(0, min(D_user[0][hour_i], min( tmp_s_on_t[0][hour_i], Rate_s)))
                tmp_E_sso[0][hour_i] = random.uniform(0, min(max( tmp_s_on_t[0][hour_i] - tmp_E_uus[0][hour_i], 0), Rate_s))
                # 電動車
                if Theta_Work[DAY][hour_i] == 0:
                    tmp_E_uuv[0][hour_i] = random.uniform(0, min(max(0, D_user[0][hour_i] - tmp_E_uus[0][hour_i]), min(tmp_ev_on_t[0][hour_i] - E_ev_min, Rate_ev)))
                    tmp_E_vso[0][hour_i] = random.uniform(0, min(max(tmp_ev_on_t[0][hour_i] - tmp_E_uuv[0][hour_i] - E_ev_min, 0), Rate_ev))
            else:
                print('Error in initial!!')
            
            # 隨機產生E_uug
            tmp_E_uug[0][hour_i] = random.uniform(0, max(D_user[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_uus[0][hour_i], 0))
            
            # E_ubo必須用來滿足D_user
            tmp_E_ubo[0][hour_i] = max(D_user[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_uus[0][hour_i] - tmp_E_uug[0][hour_i], 0)
            tmp_s_on_t[0][hour_i+1] = tmp_s_on_t[0][hour_i] + tmp_E_sfg[0][hour_i] + tmp_E_sfo[0][hour_i] - tmp_E_uus[0][hour_i] - tmp_E_sso[0][hour_i] - tmp_E_vfs[0][hour_i] + ParaOfRenergy[0][hour_i]
            tmp_ev_on_t[0][hour_i+1] = tmp_ev_on_t[0][hour_i] + tmp_E_vfg[0][hour_i] + tmp_E_vfo[0][hour_i] - tmp_E_uuv[0][hour_i] - tmp_E_vso[0][hour_i]
            
            # 將產生候選解加到人口中
            population[i].E_uug.append(tmp_E_uug[0][hour_i])
            population[i].E_sfg.append(tmp_E_sfg[0][hour_i])
            population[i].E_uus.append(tmp_E_uus[0][hour_i])
            population[i].E_sso.append(tmp_E_sso[0][hour_i])
            population[i].E_sfo.append(tmp_E_sfo[0][hour_i])
            population[i].E_ubo.append(tmp_E_ubo[0][hour_i])
            population[i].E_vfg.append(tmp_E_vfg[0][hour_i])
            population[i].E_uuv.append(tmp_E_uuv[0][hour_i])
            population[i].E_vfs.append(tmp_E_vfs[0][hour_i])
            population[i].E_vso.append(tmp_E_vso[0][hour_i])
            population[i].E_vfo.append(tmp_E_vfo[0][hour_i])
            population[i].theta_s.append(tmp_theta_s[0][hour_i])
            population[i].theta_ev.append(tmp_theta_ev[0][hour_i])
            population[i].check_s_on_t.append( tmp_s_on_t[0][hour_i +1])
            population[i].check_ev_on_t.append( tmp_ev_on_t[0][hour_i +1])
        
    # Step1.5:檢查是否有不可行解    
    for i in range( NumOfPerson):
        ChechLegalOrNot( population[i])
        
    # Step2:計算適應度值    
    for i in range( NumOfPerson):
        population[i].fitness = calTheFit( population[i])
        
    # Step2.5:存取t_0世代下最好與最差的解
    
    FitnessArray = list()                                   # 用來存當世代所有候選解的fitness
    for i in range( NumOfPerson):
        FitnessArray.append(population[i].fitness)          # temp矩陣存當世代所有人的fitness
        
    IndexOfBestCS = FitnessArray.index( min(FitnessArray))              # 找到第0世代最好的解 並返回索引值
    IndexOfWorstCS = FitnessArray.index( max( FitnessArray))            # 找到第0世代最差的解 並返回索引值
    
    
    Hist_BestFit.append( population[IndexOfBestCS].fitness)
    Hist_AvgFit.append( sum( FitnessArray)/len( FitnessArray))
    #Hist_forCompare.append( population[IndexOfBestCS].fitness)
    
    # Step3:主迴圈開始    
    for iteration in range( MaxOfIteration):
        
        
        time_iter_start = time.time()
        
        
        OffSpring = list()                                  # 宣告子代
        for i in range(1):                                  # 1次產生1條子代
            OffSpring.append(person())                      
            
            # 初始化儲電設備與電動車每小時的電量 
            tmp_E_uug, tmp_E_sfg, tmp_E_uus = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
            tmp_E_sso, tmp_E_sfo, tmp_E_ubo = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
            tmp_E_vfg, tmp_E_uuv, tmp_E_vfs = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
            tmp_E_vso, tmp_E_vfo, tmp_theta_s, tmp_theta_ev = np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours)), np.zeros((1, Hours))
            tmp_s_on_t, tmp_ev_on_t = np.zeros((1, Hours + 1)), np.zeros((1, Hours + 1))
        
            
            if DAY == 0:
                tmp_s_on_t[0][0] = E_s_min
                tmp_ev_on_t[0][0] = E_ev_min
                
            else:
                file = load_workbook( PathOfFile + 'Remain0416.xlsx')
                sheet_ranges = file['HowManyStillHave_HSA']
                
                for k in range( len( DAY_List)):
                    if DAY == list(enumerate(DAY_List))[k][0]:
                        tmp_s_on_t[0][0] = sheet_ranges.cell( row = k + 1, column = 2).value
                        tmp_ev_on_t[0][0] = sheet_ranges.cell( row = k + 1, column = 3).value
                        break
                    
            OffSpring[0].check_s_on_t.append( tmp_s_on_t[0][0])
            OffSpring[0].check_ev_on_t.append( tmp_ev_on_t[0][0])
    
            # 每欄
            for hour_i in range( Hours):
                tmp_hmcr = np.random.random()
                
                # 找舊解來改
                if tmp_hmcr < Par_HMCR:
                    
                    tmp_indexOfparent = np.random.randint(0, NumOfPerson)       # [0, NumOfPerson-1]
                
                    # theta_s
                    tmp_theta_s[0][hour_i] = population[tmp_indexOfparent].theta_s[hour_i]
                    # theta_ev
                    if hour_i > 6 and hour_i < 17:                          # 上班時段
                        tmp_theta_ev[0][hour_i] = 0
                    else:                                                   # 非上班時段
                        tmp_theta_ev[0][hour_i] = population[tmp_indexOfparent].theta_ev[hour_i]
                    
                    # 是否要微調
                    tmp_par = np.random.random()
                    
                    # 要微調
                    if tmp_par < Par_PAR:
                        tmp_E_uug[0][hour_i] = max( population[tmp_indexOfparent].E_uug[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_sfg[0][hour_i] = max( population[tmp_indexOfparent].E_sfg[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_uus[0][hour_i] = max( population[tmp_indexOfparent].E_uus[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_sso[0][hour_i] = max( population[tmp_indexOfparent].E_sso[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_sfo[0][hour_i] = max( population[tmp_indexOfparent].E_sfo[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_ubo[0][hour_i] = max( population[tmp_indexOfparent].E_ubo[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_vfg[0][hour_i] = max( population[tmp_indexOfparent].E_vfg[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_uuv[0][hour_i] = max( population[tmp_indexOfparent].E_uuv[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_vfs[0][hour_i] = max( population[tmp_indexOfparent].E_vfs[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_vso[0][hour_i] = max( population[tmp_indexOfparent].E_vso[hour_i] + np.random.uniform(-Interval, Interval), 0)
                        tmp_E_vfo[0][hour_i] = max( population[tmp_indexOfparent].E_vfo[hour_i] + np.random.uniform(-Interval, Interval), 0)
                    
                    # 維持舊解
                    else:
                        tmp_E_uug[0][hour_i] = population[tmp_indexOfparent].E_uug[hour_i]
                        tmp_E_sfg[0][hour_i] = population[tmp_indexOfparent].E_sfg[hour_i]
                        tmp_E_uus[0][hour_i] = population[tmp_indexOfparent].E_uus[hour_i]
                        tmp_E_sso[0][hour_i] = population[tmp_indexOfparent].E_sso[hour_i]
                        tmp_E_sfo[0][hour_i] = population[tmp_indexOfparent].E_sfo[hour_i]
                        tmp_E_ubo[0][hour_i] = population[tmp_indexOfparent].E_ubo[hour_i]
                        tmp_E_vfg[0][hour_i] = population[tmp_indexOfparent].E_vfg[hour_i]
                        tmp_E_uuv[0][hour_i] = population[tmp_indexOfparent].E_uuv[hour_i]
                        tmp_E_vfs[0][hour_i] = population[tmp_indexOfparent].E_vfs[hour_i]
                        tmp_E_vso[0][hour_i] = population[tmp_indexOfparent].E_vso[hour_i]
                        tmp_E_vfo[0][hour_i] = population[tmp_indexOfparent].E_vfo[hour_i]
                        
                # 範圍內隨機產生新解
                else:
                    # theta_s
                    tmp_theta_s[0][hour_i] = np.random.randint(0, 2)
                    # theta_ev
                    if hour_i > 6 and hour_i < 17:                          # 上班時段
                        tmp_theta_ev[0][hour_i] = 0
                    else:                                                   # 非上班時段
                        tmp_theta_ev[0][hour_i] = np.random.randint(0, 2)
            
                    tmp_E_uug[0][hour_i] = np.random.uniform( 0, Rate_ev)
                    tmp_E_sfg[0][hour_i] = np.random.uniform( 0, Rate_s)
                    tmp_E_uus[0][hour_i] = np.random.uniform( 0, Rate_s)
                    tmp_E_sso[0][hour_i] = np.random.uniform( 0, Rate_s)
                    tmp_E_sfo[0][hour_i] = np.random.uniform( 0, Rate_s)
                    tmp_E_ubo[0][hour_i] = np.random.uniform( 0, Rate_ev)
                    tmp_E_vfg[0][hour_i] = np.random.uniform( 0, Rate_ev)
                    tmp_E_uuv[0][hour_i] = np.random.uniform( 0, Rate_ev)
                    tmp_E_vfs[0][hour_i] = np.random.uniform( 0, Rate_ev)
                    tmp_E_vso[0][hour_i] = np.random.uniform( 0, Rate_ev)
                    tmp_E_vfo[0][hour_i] = np.random.uniform( 0, Rate_ev)
        
                OffSpring[0].E_uug.append(tmp_E_uug[0][hour_i])
                OffSpring[0].E_sfg.append(tmp_E_sfg[0][hour_i])
                OffSpring[0].E_uus.append(tmp_E_uus[0][hour_i])
                OffSpring[0].E_sso.append(tmp_E_sso[0][hour_i])
                OffSpring[0].E_sfo.append(tmp_E_sfo[0][hour_i])
                OffSpring[0].E_ubo.append(tmp_E_ubo[0][hour_i])
                OffSpring[0].E_vfg.append(tmp_E_vfg[0][hour_i])
                OffSpring[0].E_uuv.append(tmp_E_uuv[0][hour_i])
                OffSpring[0].E_vfs.append(tmp_E_vfs[0][hour_i])
                OffSpring[0].E_vso.append(tmp_E_vso[0][hour_i])
                OffSpring[0].E_vfo.append(tmp_E_vfo[0][hour_i])
                OffSpring[0].theta_s.append(tmp_theta_s[0][hour_i])
                OffSpring[0].theta_ev.append(tmp_theta_ev[0][hour_i])
            
            NPChech = np.concatenate( ( tmp_E_uug, tmp_E_sfg, tmp_E_uus, tmp_E_sso, 
                                        tmp_E_sfo, tmp_E_ubo, tmp_E_vfg, tmp_E_uuv, 
                                        tmp_E_vfs, tmp_E_vso, tmp_E_vfo, tmp_theta_s, 
                                        tmp_theta_ev))
            
        # Step4:修正不可行解
        for i in range( 1):
            
            OffSpring[i].check_s_on_t = list()
            OffSpring[i].check_ev_on_t = list()
            
            if DAY == 0:
                OffSpring[i].check_s_on_t.append(E_s_min)
                OffSpring[i].check_ev_on_t.append(E_ev_min) 
            else:
                file = load_workbook( PathOfFile + 'Remain0416.xlsx')
                sheet_ranges = file['HowManyStillHave_HSA']
                
                for k in range( len( DAY_List)):
                    if DAY == list(enumerate(DAY_List))[k][0]:
                        OffSpring[i].check_s_on_t.append( sheet_ranges.cell( row = k + 1, column = 2).value)
                        OffSpring[i].check_ev_on_t.append( sheet_ranges.cell( row = k + 1, column = 3).value) 
                        break
                            
            for hour_i in range( Hours):
                OffSpring[i].check_s_on_t.append( 0)
                OffSpring[i].check_ev_on_t.append( 0)
            
            for hour_i in range( Hours):
                
                # 1. 檢查儲放電設備是否有放電錯誤
                if OffSpring[i].theta_s[hour_i] == 0:
                    if OffSpring[i].E_sfg[hour_i] > 0:
                        OffSpring[i].E_sfg[hour_i] = 0
    
                    if OffSpring[i].E_sfo[hour_i] > 0:
                        OffSpring[i].E_sfo[hour_i] = 0
                else:
                    if OffSpring[i].E_uus[hour_i] > 0:
                        OffSpring[i].E_uug[hour_i] = OffSpring[i].E_uug[hour_i] + OffSpring[i].E_uus[hour_i]
                        OffSpring[i].E_uus[hour_i] = 0
                    
                    if OffSpring[i].E_sso[hour_i] > 0:
                        OffSpring[i].E_sso[hour_i] = 0
                    
                    if OffSpring[i].E_vfs[hour_i] > 0:
                        OffSpring[i].E_vfg[hour_i] = OffSpring[i].E_vfg[hour_i] + OffSpring[i].E_vfs[hour_i]
                        OffSpring[i].E_vfs[hour_i] = 0
                
                # 2. 檢查電動車是否由儲放電錯誤
                # 如果在上班時段
                if Theta_Work[DAY][hour_i] == 1:
                    if OffSpring[i].E_vfg[hour_i] > 0:
                        OffSpring[i].E_vfg[hour_i] = 0
                    if OffSpring[i].E_vfs[hour_i] > 0:
                        OffSpring[i].E_sso[hour_i] = OffSpring[i].E_sso[hour_i] + OffSpring[i].E_vfs[hour_i]
                        OffSpring[i].E_vfs[hour_i] = 0
                    if OffSpring[i].E_vfo[hour_i] > 0:
                        OffSpring[i].E_vfo[hour_i] = 0
                    if OffSpring[i].E_uuv[hour_i] > 0:
                        OffSpring[i].E_uug[hour_i] = OffSpring[i].E_uuv[hour_i] + OffSpring[i].E_uuv[hour_i]
                        OffSpring[i].E_uuv[hour_i] = 0
                    if OffSpring[i].E_vso[hour_i] > 0:
                        OffSpring[i].E_vso[hour_i] = 0
        
                if OffSpring[i].theta_ev[hour_i] == 0:
                    if OffSpring[i].E_vfg[hour_i] > 0:
                        OffSpring[i].E_vfg[hour_i] = 0
                    if OffSpring[i].E_vfs[hour_i] > 0:
                        OffSpring[i].E_sso[hour_i] = OffSpring[i].E_sso[hour_i] + OffSpring[i].E_vfs[hour_i]
                        OffSpring[i].E_vfs[hour_i] = 0
                    if OffSpring[i].E_vfo[hour_i] > 0:
                        OffSpring[i].E_vfo[hour_i] = 0
                else:
                    if OffSpring[i].E_uuv[hour_i] > 0:
                        OffSpring[i].E_uug[hour_i] = OffSpring[i].E_uug[hour_i] + OffSpring[i].E_uuv[hour_i]
                        OffSpring[i].E_uuv[hour_i] = 0
                    if OffSpring[i].E_vso[hour_i] > 0:
                        OffSpring[i].E_vso[hour_i] = 0
                
                # 3. 檢查儲電設備與電動車容量限制
                tmp_s = OffSpring[i].check_s_on_t[hour_i] + OffSpring[i].E_sfg[hour_i] + OffSpring[i].E_sfo[hour_i] - OffSpring[i].E_uus[hour_i] - OffSpring[i].E_sso[hour_i] - OffSpring[i].E_vfs[hour_i] + ParaOfRenergy[0][hour_i]
                tmp_ev = OffSpring[i].check_ev_on_t[hour_i] + OffSpring[i].E_vfg[hour_i] + OffSpring[i].E_vfs[hour_i] + OffSpring[i].E_vfo[hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_vso[hour_i]
                
                
                # 當儲電設備發生沒電還放電或是滿載還儲電的話, 就要修正
                if ((OffSpring[i].check_s_on_t[ hour_i + 1] < E_s_min) or (OffSpring[i].check_s_on_t[ hour_i + 1] > E_s_max)) or ((OffSpring[i].check_ev_on_t[ hour_i + 1] < E_ev_min) or ( OffSpring[i].check_ev_on_t[ hour_i + 1] > E_ev_max)):
                                             
                    OffSpring[i].E_uug[hour_i] = 0
                    OffSpring[i].E_sfg[hour_i] = 0
                    OffSpring[i].E_uus[hour_i] = 0
                    OffSpring[i].E_sso[hour_i] = 0
                    OffSpring[i].E_sfo[hour_i] = 0
                    OffSpring[i].E_ubo[hour_i] = 0
                    OffSpring[i].E_vfg[hour_i] = 0
                    OffSpring[i].E_uuv[hour_i] = 0
                    OffSpring[i].E_vfs[hour_i] = 0
                    OffSpring[i].E_vso[hour_i] = 0
                    OffSpring[i].E_vfo[hour_i] = 0
                    
                    # 儲電設備沒存量不能放電
                    if OffSpring[i].check_s_on_t[ hour_i] <= E_s_min:
                        OffSpring[i].theta_s[hour_i] = 1
                    # 儲電設備沒容量不能儲電
                    elif OffSpring[i].check_s_on_t[ hour_i] + ParaOfRenergy[0][hour_i] >= E_s_max:
                        OffSpring[i].theta_s[hour_i] = 0
                    else:
                        OffSpring[i].theta_s[hour_i] = np.random.randint(0, 2)
                
                    # 電動車沒存量不能放電
                    if OffSpring[i].check_ev_on_t[ hour_i] <= E_ev_min:
                        OffSpring[i].theta_ev[hour_i] = 1
                    # 電動車沒容量不能儲電
                    elif OffSpring[i].check_ev_on_t[ hour_i] >= E_ev_max:
                        OffSpring[i].theta_ev[hour_i] = 0
                    else:
                        OffSpring[i].theta_ev[hour_i] = np.random.randint(0, 2)
                    
                    # 儲電設備及電動車同時儲電情況
                    if OffSpring[i].theta_s[hour_i] == 1 and OffSpring[i].theta_ev[hour_i] == 1:
                        # 儲電設備
                        OffSpring[i].E_sfg[hour_i] = random.uniform( 0, min( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - ParaOfRenergy[0][hour_i], Rate_s) )
                        OffSpring[i].E_sfo[hour_i] = random.uniform( 0, min( max( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_sfg[hour_i] - ParaOfRenergy[0][hour_i], 0), Rate_s ))
                        # 電動車
                        if Theta_Work[DAY][hour_i] == 0:
                            OffSpring[i].E_vfg[hour_i] = random.uniform( 0, min( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i], Rate_ev))
                            OffSpring[i].E_vfo[hour_i] = random.uniform( 0, min( max( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_vfg[hour_i], 0), Rate_ev))
                    
                    # 產生儲放電值，當儲電設備放電，電動車儲電
                    elif OffSpring[i].theta_s[hour_i] == 0 and OffSpring[i].theta_ev[hour_i] == 1:
                        # 儲電設備
                        OffSpring[i].E_uus[hour_i] = random.uniform( 0, min( D_user[0][hour_i], min( OffSpring[i].check_s_on_t[ hour_i], Rate_s)))
                        OffSpring[i].E_sso[hour_i] = random.uniform( 0, min( max( OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_uus[hour_i], 0), Rate_s))
                        # 電動車
                        if Theta_Work[DAY][hour_i] == 0:
                            OffSpring[i].E_vfs[hour_i] = random.uniform( 0, min( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i], min( max(  OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_sso[hour_i] - OffSpring[i].E_uus[hour_i],0), Rate_s)))
                            OffSpring[i].E_vfg[hour_i] = random.uniform( 0, min( max( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_vfs[hour_i], 0), Rate_ev))
                            OffSpring[i].E_vfo[hour_i] = random.uniform( 0, min( max( E_ev_max - OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_vfg[hour_i] - OffSpring[i].E_vfs[hour_i], 0), Rate_ev))
                    
                    # 產生儲放電值，當儲電設備儲電，電動車放電
                    elif OffSpring[i].theta_s[hour_i] == 1 and OffSpring[i].theta_ev[hour_i] == 0:
                        # 儲電設備
                        OffSpring[i].E_sfg[hour_i] = random.uniform(0, min( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - ParaOfRenergy[0][hour_i], Rate_s))
                        OffSpring[i].E_sfo[hour_i] = random.uniform( 0, min( max( E_s_max - OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_sfg[hour_i] - ParaOfRenergy[0][hour_i], 0), Rate_s ))
                        # 電動車
                        if Theta_Work[DAY][hour_i] == 0:
                            OffSpring[i].E_uuv[hour_i] = random.uniform(0, min( D_user[0][hour_i], min( OffSpring[i].check_ev_on_t[ hour_i] - E_ev_min, Rate_ev)))
                            OffSpring[i].E_vso[hour_i] = random.uniform(0, min(max( OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_uuv[hour_i] - E_ev_min, 0), Rate_ev))
                    
                    # 產生儲放電值，當儲電設備放電，電動車放電
                    else:
                        # 儲電設備
                        OffSpring[i].E_uus[hour_i] = random.uniform( 0, min( D_user[0][hour_i], min( OffSpring[i].check_s_on_t[ hour_i], Rate_s)))
                        OffSpring[i].E_sso[hour_i] = random.uniform( 0, min( max( OffSpring[i].check_s_on_t[ hour_i] - OffSpring[i].E_uus[hour_i], 0), Rate_s))
                        # 電動車
                        if Theta_Work[DAY][hour_i] == 0:
                            OffSpring[i].E_uuv[hour_i] = random.uniform( 0, min( max( 0, D_user[0][hour_i] - OffSpring[i].E_uus[hour_i]), min( OffSpring[i].check_ev_on_t[ hour_i] - E_ev_min, Rate_ev)))
                            OffSpring[i].E_vso[hour_i] = random.uniform( 0, min( max( OffSpring[i].check_ev_on_t[ hour_i] - OffSpring[i].E_uuv[hour_i] - E_ev_min, 0), Rate_ev))
                    
                    # 隨機產生E_uug
                    OffSpring[i].E_uug[hour_i] = random.uniform( 0, max( D_user[0][hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_uus[hour_i], 0))
                    # E_ubo必須用來滿足D_user
                    OffSpring[i].E_ubo[hour_i] = max( D_user[0][hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_uus[hour_i] - OffSpring[i].E_uug[hour_i], 0)
                        
                # 更新結果
                tmp_s = OffSpring[i].check_s_on_t[hour_i] + OffSpring[i].E_sfg[hour_i] + OffSpring[i].E_sfo[hour_i] - OffSpring[i].E_uus[hour_i] - OffSpring[i].E_sso[hour_i] - OffSpring[i].E_vfs[hour_i] + ParaOfRenergy[0][hour_i]
                tmp_ev = OffSpring[i].check_ev_on_t[hour_i] + OffSpring[i].E_vfg[hour_i] + OffSpring[i].E_vfs[hour_i] + OffSpring[i].E_vfo[hour_i] - OffSpring[i].E_uuv[hour_i] - OffSpring[i].E_vso[hour_i]
                            
                OffSpring[i].check_s_on_t[hour_i +1] = tmp_s
                OffSpring[i].check_ev_on_t[hour_i +1] = tmp_ev
                        
        # Step5:檢查修正後是否有有不可行解        
        for i in range(1):
            ChechLegalOrNot( population[i])
            
        # Step6:計算適應度值    
        for i in range(1):
            OffSpring[i].fitness = calTheFit( OffSpring[i])            
    
        # Step7:取代最差的解
        # 若有比現有解來的優的話, 取代之
        if OffSpring[0].fitness < max( FitnessArray):
            del population[IndexOfWorstCS]
            population.append( OffSpring[0])
        
        # 將最佳解與最差解存起來
        FitnessArray = list()                                   # 用來存當世代所有候選解的fitness
        for i in range( NumOfPerson):
            FitnessArray.append(population[i].fitness)          # temp矩陣存當世代所有人的fitness
            
        IndexOfBestCS = FitnessArray.index( min(FitnessArray))              # 找到第0世代最好的解 並返回索引值
        IndexOfWorstCS = FitnessArray.index( max( FitnessArray))            # 找到第0世代最差的解 並返回索引值
        
        Hist_BestFit.append( population[IndexOfBestCS].fitness)
        Hist_AvgFit.append( sum( FitnessArray)/len( FitnessArray))
        
    #    if iteration % 20 == 0:
    #        Hist_forCompare.append( Hist_BestFit[-1])
        
        time_iter_end = time.time()
#        
        time_FLAG = time.time()
        
        if iteration % 5000 == 0:
            print( 'now in', iteration, 'iteration')
    #        print( 'one iteration spent:', time_iter_end - time_iter_start )
            print( "now are spent :", round(time_FLAG - time_start, 4))
            
#        if (( time_FLAG - time_start) >= 825):
#            break
            
    # ==== 演算法結束 ====
    # Step8:看最後的解
    BestCSSolution = np.zeros(( 13, Hours))
    for i in range( Hours):
        BestCSSolution[0][i] = population[IndexOfBestCS].E_uug[i]
        BestCSSolution[1][i] = population[IndexOfBestCS].E_sfg[i]
        BestCSSolution[2][i] = population[IndexOfBestCS].E_uus[i]
        BestCSSolution[3][i] = population[IndexOfBestCS].E_sso[i]
        BestCSSolution[4][i] = population[IndexOfBestCS].E_sfo[i]
        BestCSSolution[5][i] = population[IndexOfBestCS].E_ubo[i]
        BestCSSolution[6][i] = population[IndexOfBestCS].E_vfg[i]
        BestCSSolution[7][i] = population[IndexOfBestCS].E_uuv[i]
        BestCSSolution[8][i] = population[IndexOfBestCS].E_vfs[i]
        BestCSSolution[9][i] = population[IndexOfBestCS].E_vso[i]
        BestCSSolution[10][i] = population[IndexOfBestCS].E_vfo[i]
        BestCSSolution[11][i] = population[IndexOfBestCS].theta_s[i]
        BestCSSolution[12][i] = population[IndexOfBestCS].theta_ev[i]
    
    # 將儲電設備與電動車最後的電量儲存起來
    ToTxt_s = population[IndexOfBestCS].check_s_on_t[-1]
    ToTxt_ev = population[IndexOfBestCS].check_ev_on_t[-1]
    
    file = load_workbook( PathOfFile + 'Remain0416.xlsx')
    sheet_ranges = file['HowManyStillHave_HSA']
    
    for k in range( len( DAY_List)):
        if DAY == list(enumerate(DAY_List))[k][0]:
            sheet_ranges.cell(row = k + 2, column = 2).value = ToTxt_s
            sheet_ranges.cell(row = k + 2, column = 3).value = ToTxt_ev
            break
    
    file.save( 'Remain0416.xlsx')
    file.close()
    time_end = time.time()
    print( "time:", time_end - time_start)
    
    ####
    List_bestValue.append( Hist_BestFit[-1])
    List_time.append( time_end - time_start)
    print( "Finish", iRun + 1, "time")
    print( "#### #### #### ####")
    